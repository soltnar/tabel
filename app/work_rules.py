from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from functools import lru_cache
from io import BytesIO
from typing import Any, Optional
import calendar
import re

import pandas as pd
from openpyxl import load_workbook


WORK_CODES = {"я", "пл"}
VACATION_CODES = {"от", "од"}
SICK_CODES = {"б", "т"}


def normalize_text(value: Any) -> str:
    if value is None or (not isinstance(value, str) and pd.isna(value)):
        return ""
    text = str(value).strip().lower().replace("ё", "е")
    return re.sub(r"\s+", " ", text)


def normalize_name(value: Any) -> str:
    text = normalize_text(value)
    text = re.sub(r"\(\s*\d+\s*\)\s*$", "", text)
    return re.sub(r"[^а-яa-z]+", " ", text).strip()


def names_compatible(left: Any, right: Any) -> bool:
    """Match a full Russian name with the same name written using initials."""
    left_parts = normalize_name(left).split()
    right_parts = normalize_name(right).split()
    if not left_parts or not right_parts or left_parts[0] != right_parts[0]:
        return False
    if len(left_parts) != len(right_parts):
        return False

    return all(
        left_part == right_part
        or left_part.startswith(right_part)
        or right_part.startswith(left_part)
        for left_part, right_part in zip(left_parts[1:], right_parts[1:])
    )


def find_unique_name_match(mapping: dict[str, Any], employee: Any) -> Any:
    """Return a record only when the normalized or initials-based match is unique."""
    employee_key = normalize_name(employee)
    if employee_key in mapping:
        return mapping[employee_key]

    matches = [
        value
        for candidate, value in mapping.items()
        if names_compatible(employee_key, candidate)
    ]
    return matches[0] if len(matches) == 1 else None


def normalize_tab(value: Any) -> str:
    if value is None or (not isinstance(value, str) and pd.isna(value)):
        return ""
    text = str(value).strip()
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        return str(int(float(text)))
    match = re.search(r"\d+", text)
    return str(int(match.group())) if match else ""


def parse_date(value: Any) -> Optional[date]:
    if value is None or (not isinstance(value, str) and pd.isna(value)):
        return None
    if isinstance(value, pd.Timestamp):
        return value.date()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
    return parsed.date() if pd.notna(parsed) else None


@lru_cache(maxsize=8)
def _read_excel_any(file_bytes: bytes, filename: Optional[str], header: Optional[int] = None):
    suffix = str(filename or "").lower()
    engine = "xlrd" if suffix.endswith(".xls") and not suffix.endswith(".xlsx") else None
    return pd.read_excel(BytesIO(file_bytes), sheet_name=None, header=header, dtype=object, engine=engine)


def parse_personnel_events(
    file_bytes: bytes,
    filename: Optional[str] = None,
) -> dict[str, list[tuple[Optional[date], Optional[date]]]]:
    workbook = _read_excel_any(file_bytes, filename, header=None)
    events: dict[str, list[tuple[date, str]]] = defaultdict(list)

    for df in workbook.values():
        header_row = None
        date_col = type_col = employee_col = None
        for row_idx in range(min(30, len(df))):
            cells = [normalize_text(v) for v in df.iloc[row_idx].tolist()]
            for col_idx, cell in enumerate(cells):
                if "дата" in cell and date_col is None:
                    date_col = col_idx
                if "тип документ" in cell and type_col is None:
                    type_col = col_idx
                if "сотрудник" in cell and employee_col is None:
                    employee_col = col_idx
            if date_col is not None and type_col is not None and employee_col is not None:
                header_row = row_idx
                break
        if header_row is None:
            continue

        for row_idx in range(header_row + 1, len(df)):
            event_date = parse_date(df.iat[row_idx, date_col])
            event_type = normalize_text(df.iat[row_idx, type_col])
            employee = normalize_name(df.iat[row_idx, employee_col])
            if event_date and employee and ("прием" in event_type or "уволь" in event_type):
                events[employee].append((event_date, "hire" if "прием" in event_type else "dismiss"))

    intervals: dict[str, list[tuple[Optional[date], Optional[date]]]] = {}
    for employee, employee_events in events.items():
        active_start: Optional[date] = None
        employee_intervals: list[tuple[Optional[date], Optional[date]]] = []
        for event_date, event_type in sorted(set(employee_events), key=lambda item: (item[0], item[1])):
            if event_type == "hire":
                if active_start is None:
                    active_start = event_date
            elif active_start is not None:
                employee_intervals.append((active_start, event_date))
                active_start = None
            else:
                employee_intervals.append((None, event_date))
        if active_start is not None:
            employee_intervals.append((active_start, None))
        intervals[employee] = employee_intervals
    return intervals


def _find_register_sections(df: pd.DataFrame) -> list[tuple[int, int]]:
    starts: list[int] = []
    for row_idx in range(len(df)):
        row_text = " ".join(normalize_text(v) for v in df.iloc[row_idx].tolist() if normalize_text(v))
        if "расчетная ведомость" in row_text:
            starts.append(row_idx)
    return [(start, starts[pos + 1] if pos + 1 < len(starts) else len(df)) for pos, start in enumerate(starts)]


def _find_header_cell(df: pd.DataFrame, start: int, end: int, tokens: tuple[str, ...]) -> Optional[tuple[int, int]]:
    for row_idx in range(start, min(end, start + 25)):
        for col_idx, value in enumerate(df.iloc[row_idx].tolist()):
            text = normalize_text(value)
            if text and all(token in text for token in tokens):
                return row_idx, col_idx
    return None


def parse_annual_payroll_register(
    file_bytes: bytes,
    year: int,
    month: int,
    filename: Optional[str] = None,
) -> dict[str, dict[str, Any]]:
    workbook = _read_excel_any(file_bytes, filename, header=None)
    target: dict[str, dict[str, Any]] = {}
    month_start = date(year, month, 1)
    month_end = date(year, month, calendar.monthrange(year, month)[1])
    organization = ""

    for df in workbook.values():
        for row_idx in range(min(20, len(df))):
            for value in df.iloc[row_idx].tolist():
                text = str(value or "").strip()
                if re.search(r"\b(?:ооо|ао|пао|ип)\b", normalize_text(text)):
                    organization = text
                    break
            if organization:
                break
        if organization:
            break

    for df in workbook.values():
        sections = _find_register_sections(df)
        if not sections:
            sections = [(0, len(df))]
        for section_index, (start, end) in enumerate(sections, start=1):
            if len(sections) >= 12 and section_index != month:
                continue

            tab_header = _find_header_cell(df, start, end, ("табель",))
            name_header = _find_header_cell(df, start, end, ("фио",))
            if name_header is None:
                name_header = _find_header_cell(df, start, end, ("фамил",))
            days_header = _find_header_cell(df, start, end, ("рабоч",))
            birth_header = _find_header_cell(df, start, end, ("дата", "рожд"))
            adulthood_header = _find_header_cell(df, start, end, ("совершеннолет",))
            sick_header = _find_header_cell(df, start, end, ("больнич",))
            vacation_header = _find_header_cell(df, start, end, ("отпуск",))
            role_header = _find_header_cell(df, start, end, ("должност",))

            # В ведомостях пользователя стабильны колонки табельного номера и ФИО.
            tab_col = tab_header[1] if tab_header else 2
            name_col = name_header[1] if name_header else 4
            work_col = days_header[1] if days_header else 13
            birth_col = adulthood_header[1] if adulthood_header else (birth_header[1] if birth_header else None)
            sick_col = sick_header[1] if sick_header else 26
            vacation_col = vacation_header[1] if vacation_header else 28
            role_col = role_header[1] if role_header else 6

            data_start = max(
                [cell[0] for cell in (tab_header, name_header, days_header) if cell is not None] or [start]
            ) + 1
            for row_idx in range(data_start, end):
                tab = normalize_tab(df.iat[row_idx, tab_col] if tab_col < df.shape[1] else None)
                employee = normalize_name(df.iat[row_idx, name_col] if name_col < df.shape[1] else None)
                if not tab and not employee:
                    continue
                if not employee or len(employee.split()) < 2:
                    continue

                rec = target.setdefault(
                    tab or employee,
                    {
                        "tab_number": tab,
                        "employee": employee,
                        "vacation_days": set(),
                        "sick_days": set(),
                        "register_days": None,
                        "register_hours": None,
                        "birth_date": None,
                        "adulthood_date": None,
                        "role": "",
                        "organization": organization,
                    },
                )

                role = normalize_text(df.iat[row_idx, role_col] if role_col < df.shape[1] else None)
                if role:
                    rec["role"] = role

                work_value = df.iat[row_idx, work_col] if work_col < df.shape[1] else None
                work_match = re.search(r"(\d+)\s*\(\s*([\d\s,.]+)\s*\)", str(work_value or ""))
                if work_match:
                    rec["register_days"] = int(work_match.group(1))
                    rec["register_hours"] = float(work_match.group(2).replace(" ", "").replace(",", "."))

                if birth_col is not None and birth_col < df.shape[1]:
                    dt = parse_date(df.iat[row_idx, birth_col])
                    if dt:
                        if "совершеннолет" in normalize_text(
                            df.iat[adulthood_header[0], adulthood_header[1]] if adulthood_header else ""
                        ) or dt > month_end:
                            rec["adulthood_date"] = dt
                            try:
                                rec["birth_date"] = date(dt.year - 18, dt.month, dt.day)
                            except ValueError:
                                rec["birth_date"] = date(dt.year - 18, dt.month, 28)
                        elif dt.year <= year - 10:
                            rec["birth_date"] = dt

                for col, key in ((sick_col, "sick_days"), (vacation_col, "vacation_days")):
                    if col >= df.shape[1]:
                        continue
                    range_start = parse_date(df.iat[row_idx, col])
                    range_end = parse_date(df.iat[row_idx, col + 1]) if col + 1 < df.shape[1] else None
                    if range_start and not range_end:
                        range_end = range_start
                    if range_start and range_end:
                        cursor = max(range_start, month_start)
                        finish = min(range_end, month_end)
                        while cursor <= finish:
                            rec[key].add(cursor.day)
                            cursor += pd.Timedelta(days=1)
    return target


def infer_timesheet_period(
    files: list[tuple[str, bytes]],
) -> Optional[tuple[int, int]]:
    """Return a single unambiguous period found in uploaded T-13 files."""
    periods: set[tuple[int, int]] = set()
    for filename, file_bytes in files:
        try:
            wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
        except Exception:
            continue
        sheets = [wb[name] for name in wb.sheetnames if "общ" in normalize_text(name)]
        if not sheets and wb.sheetnames:
            sheets = [wb[wb.sheetnames[0]]]
        if not sheets:
            continue
        period = _infer_period_from_workbook(sheets[0], filename)
        if period:
            periods.add(period)
    return next(iter(periods)) if len(periods) == 1 else None


def _infer_period_from_workbook(ws, filename: str = "") -> Optional[tuple[int, int]]:
    votes: dict[tuple[int, int], int] = defaultdict(int)
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40), values_only=True):
        for value in row:
            dt = parse_date(value)
            if dt and 2020 <= dt.year <= 2100:
                votes[(dt.year, dt.month)] += 1
    name = normalize_text(filename)
    month_tokens = {
        "январ": 1, "феврал": 2, "март": 3, "апрел": 4, "май": 5, "июн": 6,
        "июл": 7, "август": 8, "сентябр": 9, "октябр": 10, "ноябр": 11, "декабр": 12,
    }
    year_match = re.search(r"20\d{2}", name)
    if year_match:
        for token, month in month_tokens.items():
            if token in name:
                votes[(int(year_match.group()), month)] += 5
    return max(votes, key=votes.get) if votes else None


def _find_t13_day_columns(rows: list[tuple[Any, ...]]) -> tuple[Optional[int], dict[int, int]]:
    """Find the T-13 header and its non-contiguous day columns."""
    max_scan_row = min(len(rows), 500)
    for upper_idx in range(max_scan_row):
        upper: dict[int, int] = {}
        for col_idx, value in enumerate(rows[upper_idx], start=1):
            if isinstance(value, (int, float)) and float(value).is_integer():
                day = int(value)
                if 1 <= day <= 15:
                    upper[day] = col_idx
        if len(upper) < 12:
            continue

        for lower_idx in range(upper_idx + 1, min(upper_idx + 8, max_scan_row - 1) + 1):
            lower: dict[int, int] = {}
            for col_idx, value in enumerate(rows[lower_idx], start=1):
                if isinstance(value, (int, float)) and float(value).is_integer():
                    day = int(value)
                    if 16 <= day <= 31:
                        lower[day] = col_idx
            if len(lower) >= 12:
                return lower_idx + 1, {**upper, **lower}
    return None, {}


def extract_t13_records(
    ws,
    filename: str = "",
    expected_period: Optional[tuple[int, int]] = None,
) -> tuple[Optional[tuple[int, int]], dict[str, dict[str, Any]]]:
    """Read employee day codes/hours from a four-row T-13 employee layout."""
    period = _infer_period_from_workbook(ws, filename)
    if expected_period and period and period != expected_period:
        return period, {}

    # Read-only worksheets are efficient sequentially but extremely slow for
    # repeated ws.cell() lookups. Historical forms fit well below this bound.
    rows = list(
        ws.iter_rows(
            min_row=1,
            max_row=min(ws.max_row, 5000),
            max_col=min(ws.max_column, 120),
            values_only=True,
        )
    )
    header_row, day_columns = _find_t13_day_columns(rows)
    if header_row is None or not day_columns:
        return period, {}

    records: dict[str, dict[str, Any]] = {}
    month_days = calendar.monthrange(*(expected_period or period))[1] if (expected_period or period) else 31
    start_row = header_row + 1

    def value_at(row_number: int, column_number: int):
        if row_number < 1 or row_number > len(rows):
            return None
        row = rows[row_number - 1]
        return row[column_number - 1] if column_number <= len(row) else None

    for row_idx in range(start_row, len(rows) - 2):
        tab = normalize_tab(value_at(row_idx, 4))
        employee_cell = value_at(row_idx, 2)
        employee = normalize_name(str(employee_cell or "").split("(", 1)[0])
        if not tab or not employee:
            continue

        rec = records.setdefault(
            tab,
            {
                "tab_number": tab,
                "employee": employee,
                "work_hours": {},
                "absence_codes": {},
                "day_codes": {},
                "source": filename,
            },
        )
        for day, col in day_columns.items():
            if day > month_days:
                continue
            code_row = row_idx if day <= 15 else row_idx + 2
            hours_row = code_row + 1
            code = normalize_text(value_at(code_row, col)).upper()
            hours = pd.to_numeric(value_at(hours_row, col), errors="coerce")
            if not code:
                continue
            rec["day_codes"][day] = code
            if normalize_text(code) in WORK_CODES:
                rec["work_hours"][day] = float(hours) if pd.notna(hours) else 0.0
            elif normalize_text(code) in VACATION_CODES | SICK_CODES:
                rec["absence_codes"][day] = code
    return period, records


def parse_t13_overrides(
    files: list[tuple[str, bytes]],
    year: int,
    month: int,
) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}

    for filename, file_bytes in files:
        try:
            wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
        except Exception:
            continue
        sheets = [wb[name] for name in wb.sheetnames if "общ" in normalize_text(name)]
        if not sheets:
            sheets = [wb[wb.sheetnames[0]]]
        ws = sheets[0]
        _, extracted = extract_t13_records(
            ws,
            filename=filename,
            expected_period=(year, month),
        )
        for tab, extracted_rec in extracted.items():
            rec = result.setdefault(
                tab,
                {"work_hours": {}, "absence_codes": {}, "source": filename},
            )
            rec["work_hours"].update(extracted_rec["work_hours"])
            rec["absence_codes"].update(extracted_rec["absence_codes"])
    return result


def employment_allowed_days(
    intervals: list[tuple[Optional[date], Optional[date]]],
    year: int,
    month: int,
) -> set[int]:
    if not intervals:
        return set(range(1, calendar.monthrange(year, month)[1] + 1))
    allowed: set[int] = set()
    for day in range(1, calendar.monthrange(year, month)[1] + 1):
        current = date(year, month, day)
        if any((start is None or current >= start) and (end is None or current <= end) for start, end in intervals):
            allowed.add(day)
    return allowed


def minor_daily_cap(birth_date: Optional[date], current: date) -> float:
    if birth_date is None:
        return 13.0
    age = current.year - birth_date.year - ((current.month, current.day) < (birth_date.month, birth_date.day))
    if age >= 18:
        return 13.0
    study_weekday = current.weekday() < 5 and current.month in {1, 2, 3, 4, 5, 9, 10, 11, 12}
    if age < 16:
        return 2.5 if study_weekday else 4.0
    return 4.0 if study_weekday else 7.0
