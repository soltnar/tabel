from __future__ import annotations

from collections import Counter
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from app.excel_parsers import PreparedInput
from app.work_rules import extract_t13_records, normalize_name, normalize_tab


SEVERITY_ORDER = {"CRITICAL": 0, "ERROR": 1, "WARNING": 2, "INFO": 3}


def _issue(
    issues: list[dict[str, Any]],
    severity: str,
    code: str,
    details: str,
    *,
    employee: str = "",
    tab_number: str = "",
    day: int | None = None,
    actual: Any = None,
    expected: Any = None,
    source: str = "",
) -> None:
    issues.append(
        {
            "severity": severity,
            "code": code,
            "employee": employee,
            "tab_number": tab_number,
            "day": day,
            "actual": actual,
            "expected": expected,
            "source": source,
            "details": details,
        }
    )


def _prepared_lookups(prepared: PreparedInput) -> tuple[dict[str, pd.Series], dict[str, pd.Series]]:
    by_tab: dict[str, pd.Series] = {}
    by_name: dict[str, pd.Series] = {}
    for _, row in prepared.employees.iterrows():
        tab = normalize_tab(row.get("tab_number", ""))
        name = normalize_name(row.get("employee", ""))
        if tab:
            by_tab[tab] = row
        if name:
            by_name[name] = row
    return by_tab, by_name


def _read_t13_file(
    filename: str,
    file_bytes: bytes,
) -> tuple[tuple[int, int] | None, dict[str, dict[str, Any]]]:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    sheet_names = [name for name in workbook.sheetnames if "общ" in normalize_name(name)]
    sheet = workbook[sheet_names[0] if sheet_names else workbook.sheetnames[0]]
    return extract_t13_records(sheet, filename=filename)


def validate_uploaded_timesheets(
    prepared: PreparedInput,
    timesheet_files: list[tuple[str, bytes]],
) -> dict[str, Any]:
    """Validate uploaded T-13 files without generating or exporting a schedule."""
    issues: list[dict[str, Any]] = []
    by_tab, by_name = _prepared_lookups(prepared)
    expected_period = (
        (prepared.period_year, prepared.period_month)
        if prepared.period_year is not None and prepared.period_month is not None
        else None
    )
    matched_employees: set[str] = set()
    seen_work_days: dict[str, dict[int, tuple[float, str]]] = {}
    parsed_files = 0

    for filename, file_bytes in timesheet_files:
        try:
            period, records = _read_t13_file(filename, file_bytes)
        except Exception as exc:
            _issue(
                issues,
                "CRITICAL",
                "T13_READ_ERROR",
                f"Не удалось прочитать табель: {exc}",
                source=filename,
            )
            continue

        parsed_files += 1
        if expected_period and period and period != expected_period:
            _issue(
                issues,
                "CRITICAL",
                "PERIOD_MISMATCH",
                "Отчетный период Т-13 не совпадает с расчетной ведомостью.",
                actual=f"{period[1]:02d}.{period[0]}",
                expected=f"{expected_period[1]:02d}.{expected_period[0]}",
                source=filename,
            )

        if not records:
            _issue(
                issues,
                "ERROR",
                "T13_NO_RECORDS",
                "В табеле не найдены строки сотрудников в формате Т-13.",
                source=filename,
            )
            continue

        for tab, record in records.items():
            employee_name = str(record.get("employee", "") or "")
            row = by_tab.get(normalize_tab(tab))
            if row is None:
                row = by_name.get(normalize_name(employee_name))
            if row is None:
                _issue(
                    issues,
                    "WARNING",
                    "T13_EMPLOYEE_NOT_IN_PAYROLL",
                    "Сотрудник есть в Т-13, но отсутствует в расчетной ведомости.",
                    employee=employee_name,
                    tab_number=tab,
                    source=filename,
                )
                continue

            employee = str(row.get("employee", "") or employee_name)
            employee_key = normalize_tab(row.get("tab_number", "")) or normalize_name(employee)
            matched_employees.add(employee_key)
            work_hours = {
                int(day): float(hours or 0)
                for day, hours in dict(record.get("work_hours", {})).items()
            }
            aggregate = seen_work_days.setdefault(employee_key, {})
            for day, hours in work_hours.items():
                if day in aggregate:
                    previous_hours, previous_source = aggregate[day]
                    _issue(
                        issues,
                        "ERROR",
                        "DUPLICATE_WORK_DAY",
                        f"Явка за один день найдена в нескольких табелях ({previous_source}, {filename}).",
                        employee=employee,
                        tab_number=tab,
                        day=day,
                        actual=round(previous_hours + hours, 2),
                        source=filename,
                    )
                else:
                    aggregate[day] = (hours, filename)

    for employee_key, work_days in seen_work_days.items():
        row = by_tab.get(employee_key)
        if row is None:
            row = by_name.get(employee_key)
        if row is None:
            continue

        employee = str(row.get("employee", "") or "")
        tab = normalize_tab(row.get("tab_number", ""))
        actual_days = len(work_days)
        actual_hours = round(sum(hours for hours, _ in work_days.values()), 2)
        expected_days = int(row.get("max_days", 0) or 0)
        expected_hours = round(float(row.get("max_hours", 0) or 0), 2)

        if actual_days != expected_days:
            _issue(
                issues,
                "ERROR",
                "SHIFT_TOTAL_MISMATCH",
                "Количество явок в Т-13 не равно количеству смен расчетной ведомости.",
                employee=employee,
                tab_number=tab,
                actual=actual_days,
                expected=expected_days,
            )
        if abs(actual_hours - expected_hours) > 0.02:
            _issue(
                issues,
                "ERROR",
                "HOURS_TOTAL_MISMATCH",
                "Сумма часов в Т-13 не равна часам расчетной ведомости.",
                employee=employee,
                tab_number=tab,
                actual=actual_hours,
                expected=expected_hours,
            )

        allowed_days = set(row.get("allowed_days", []) or [])
        absence_codes = dict(row.get("absence_codes", {}) or {})
        daily_caps = dict(row.get("daily_caps", {}) or {})
        for day, (hours, source) in work_days.items():
            if hours <= 0:
                _issue(
                    issues,
                    "ERROR",
                    "WORK_CODE_WITHOUT_HOURS",
                    "В Т-13 указана явка, но под ней нет положительного количества часов.",
                    employee=employee,
                    tab_number=tab,
                    day=day,
                    actual=hours,
                    source=source,
                )
            if day not in allowed_days:
                absence = absence_codes.get(day)
                if absence == "Б":
                    code = "WORK_DURING_SICK_LEAVE"
                    details = "Явка пересекается с больничным из расчетной ведомости."
                elif absence == "ОТ":
                    code = "WORK_DURING_VACATION"
                    details = "Явка пересекается с отпуском из расчетной ведомости."
                else:
                    code = "WORK_OUTSIDE_ALLOWED_PERIOD"
                    details = "Явка стоит вне разрешенного периода работы сотрудника."
                _issue(
                    issues,
                    "CRITICAL",
                    code,
                    details,
                    employee=employee,
                    tab_number=tab,
                    day=day,
                    actual=hours,
                    source=source,
                )

            cap = daily_caps.get(day)
            if cap is not None and hours > float(cap) + 0.02:
                _issue(
                    issues,
                    "CRITICAL",
                    "MINOR_DAILY_LIMIT",
                    "Смена несовершеннолетнего превышает допустимую продолжительность.",
                    employee=employee,
                    tab_number=tab,
                    day=day,
                    actual=hours,
                    expected=float(cap),
                    source=source,
                )

    for _, row in prepared.employees.iterrows():
        employee = str(row.get("employee", "") or "")
        tab = normalize_tab(row.get("tab_number", ""))
        employee_key = tab or normalize_name(employee)
        if employee_key in matched_employees:
            continue
        if int(row.get("max_days", 0) or 0) <= 0 and float(row.get("max_hours", 0) or 0) <= 0:
            continue
        _issue(
            issues,
            "ERROR",
            "PAYROLL_EMPLOYEE_NOT_IN_T13",
            "Сотрудник есть в расчетной ведомости, но не найден в загруженных Т-13.",
            employee=employee,
            tab_number=tab,
            expected=f"{int(row.get('max_days', 0) or 0)} смен / {float(row.get('max_hours', 0) or 0):g} ч",
        )

    issues.sort(
        key=lambda item: (
            SEVERITY_ORDER.get(item["severity"], 9),
            item.get("employee", ""),
            item.get("day") or 0,
            item.get("code", ""),
        )
    )
    counts = Counter(issue["severity"] for issue in issues)
    return {
        "message": (
            "Проверка завершена."
            if timesheet_files
            else "Входные данные проверены. Для сверки фактического табеля загрузите файл Т-13."
        ),
        "checked_files": parsed_files,
        "employees_checked": len(matched_employees),
        "issues_count": len(issues),
        "critical_count": counts["CRITICAL"],
        "error_count": counts["ERROR"],
        "warning_count": counts["WARNING"],
        "issues": issues,
        "input_warnings": prepared.warnings,
    }
