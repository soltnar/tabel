from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
import calendar
from io import BytesIO
from pathlib import Path
import re
from typing import Any, Optional

import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class Shift:
    key: str
    label: str
    start: str
    end: str
    hours: float


SHIFTS: tuple[Shift, ...] = (
    Shift("lunch", "бизнес-ланч", "12:00", "16:00", 4.0),
    Shift("evening", "вечерняя посадка", "18:00", "22:00", 4.0),
)

MANDATORY_SHIFT_KEYS = ("lunch", "evening")
OPTIONAL_SHIFT_KEYS: tuple[str, ...] = ()
KITCHEN_ROLE = "Кухня"
BAR_ROLE = "Бар"
HALL_ROLE = "Зал"
CASH_ROLE = "Касса"
CORE_MANDATORY_GROUPS: tuple[str, ...] = (KITCHEN_ROLE, HALL_ROLE, CASH_ROLE, BAR_ROLE)
DEFICIT_EMPLOYEE_LABEL = "НЕ ЗАКРЫТО"
MAX_CONSECUTIVE_WORK_DAYS = 6
NON_MANDATORY_ROLE_TOKENS = (
    "подсобный рабочий",
    "бухгалтер",
    "директор",
    "маркетолог",
    "аниматор",
    "гардеробщик",
)
WEEKEND_FOCUSED_OPTIONAL_ROLE_TOKENS = (
    "аниматор",
    "гардеробщик",
)
WEEKDAY_FOCUSED_OPTIONAL_ROLE_TOKENS = (
    "бухгалтер",
    "подсобный рабочий",
    "маркетолог",
    "директор",
)


class ScheduleGenerationError(RuntimeError):
    pass


@dataclass
class ScheduleResult:
    assignments: pd.DataFrame
    employee_summary: pd.DataFrame
    matrix: pd.DataFrame
    warnings: list[str]


class _EmployeeState:
    def __init__(self, row: pd.Series) -> None:
        self.employee = str(row["employee"])
        self.restaurant = str(row["restaurant"])
        self.tab_number = str(row.get("tab_number", "") or "")
        self.role_original = str(row["role_original"])
        self.role_group = str(row["role_group"])
        self.max_hours = float(row["max_hours"])
        self.max_days = int(row["max_days"])
        self.first_half_pay = float(pd.to_numeric(row.get("first_half_pay", 0.0), errors="coerce") or 0.0)
        self.second_half_pay = float(pd.to_numeric(row.get("second_half_pay", 0.0), errors="coerce") or 0.0)
        self.half_preference = str(row.get("half_preference", "neutral") or "neutral")

        self.used_hours = 0.0
        self.used_days = 0
        self.daily_hours: dict[int, float] = defaultdict(float)
        self.assigned_days: set[int] = set()
        self.daily_shift_keys: dict[int, set[str]] = defaultdict(set)
        self.daily_shift_groups: dict[int, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))

    @property
    def id(self) -> tuple[str, str, str]:
        return (self.employee, self.restaurant, self.role_group)


def _shift_by_key(key: str) -> Shift:
    for shift in SHIFTS:
        if shift.key == key:
            return shift
    raise KeyError(key)


def _is_mandatory_coverage_role(role: str) -> bool:
    role_norm = str(role).strip().lower()
    return not any(token in role_norm for token in NON_MANDATORY_ROLE_TOKENS)


def _is_core_group(role_group: str) -> bool:
    return str(role_group) in CORE_MANDATORY_GROUPS


def _required_slots_for_shift(
    day: int,
    shift_key: str,
    weekend_days: set[int],
    role: str,
    is_mandatory_role: bool,
) -> int:
    role_norm = str(role).strip().lower()
    role_group = str(role).strip()
    _is_weekend = day in weekend_days

    # Для core-групп (кухня/зал/касса/бар) нужен один сотрудник на день:
    # не дублируем обязательный набор по каждой смене.
    # Это снижает переизбыток назначений и лучше соответствует факту,
    # что итоговая выработка по часам берется из расчетных листков.
    if _is_core_group(role_group):
        return 1 if shift_key == MANDATORY_SHIFT_KEYS[0] else 0

    if not is_mandatory_role:
        if any(token in role_norm for token in WEEKEND_FOCUSED_OPTIONAL_ROLE_TOKENS):
            return 2 if _is_weekend else 0
        if any(token in role_norm for token in WEEKDAY_FOCUSED_OPTIONAL_ROLE_TOKENS):
            return 1 if not _is_weekend else 0
        return 1 if _is_weekend else 0

    base = 1
    if shift_key in MANDATORY_SHIFT_KEYS and _is_weekend:
        return base + 1
    return base


def _consecutive_span_with_day(assigned_days: set[int], day: int) -> int:
    days = set(assigned_days)
    days.add(day)

    left = day
    while (left - 1) in days:
        left -= 1

    right = day
    while (right + 1) in days:
        right += 1

    return right - left + 1


def _can_take_shift(emp: _EmployeeState, day: int, shift: Shift) -> bool:
    return _can_take_shift_for_group(emp=emp, day=day, shift=shift, role_group="")


def _is_manager_role(emp: _EmployeeState) -> bool:
    role = str(emp.role_original).strip().lower()
    return ("менеджер" in role) or ("администратор" in role) or ("адмистратор" in role)


def _is_front_multirole(emp: _EmployeeState) -> bool:
    # Универсальный фронт-персонал: менеджеры и администраторы.
    # Могут закрывать Бар/Зал/Кассу, но не Кухню.
    return _is_manager_role(emp)


def _can_share_same_shift(emp: _EmployeeState, day: int, shift_key: str, role_group: str) -> bool:
    role_group = str(role_group or "").strip()
    if role_group not in {BAR_ROLE, CASH_ROLE}:
        return False
    if not _is_manager_role(emp):
        return False

    existing_groups = emp.daily_shift_groups[day].get(shift_key, set())
    if not existing_groups:
        return False
    if role_group in existing_groups:
        return False
    return all(group in {BAR_ROLE, CASH_ROLE} for group in existing_groups)


def _can_take_shift_for_group(emp: _EmployeeState, day: int, shift: Shift, role_group: str) -> bool:
    new_day = day not in emp.assigned_days

    same_shift_exists = shift.key in emp.daily_shift_keys[day]
    shared_same_shift = same_shift_exists and _can_share_same_shift(
        emp=emp,
        day=day,
        shift_key=shift.key,
        role_group=role_group,
    )
    if same_shift_exists and not shared_same_shift:
        return False

    if not shared_same_shift:
        if emp.used_hours + shift.hours > emp.max_hours + 1e-9:
            return False

    if new_day and emp.used_days + 1 > emp.max_days:
        return False

    if not shared_same_shift:
        if emp.daily_hours[day] + shift.hours > 13.0 + 1e-9:
            return False

    if new_day:
        streak = _consecutive_span_with_day(emp.assigned_days, day)
        if streak > MAX_CONSECUTIVE_WORK_DAYS:
            return False

    return True


def _candidate_score(
    emp: _EmployeeState,
    day: int,
    prefer_existing_day: bool,
    deterministic_tiebreak: float,
    day_rank: int,
    total_days: int,
) -> float:
    hours_ratio = emp.used_hours / max(emp.max_hours, 1.0)
    days_ratio = emp.used_days / max(float(emp.max_days), 1.0)

    day_presence_penalty = 0.0
    if prefer_existing_day:
        day_presence_penalty = 0.16 if day not in emp.assigned_days else 0.0
    else:
        day_presence_penalty = 0.12 if day in emp.assigned_days else 0.0

    day_load_penalty = (emp.daily_hours[day] / 13.0) * 0.2

    projected_days = emp.used_days + (1 if day not in emp.assigned_days else 0)
    expected_days = (day_rank / max(total_days, 1)) * float(emp.max_days)
    pace_ahead = max(0.0, projected_days - expected_days)
    pace_behind = max(0.0, expected_days - projected_days)
    # Сильно штрафуем "раннее выгорание" (слишком много смен в первой половине месяца).
    pace_penalty = (0.22 * pace_ahead) + (0.05 * pace_behind)

    streak = _consecutive_span_with_day(emp.assigned_days, day)
    # После 3-4 дней подряд рейтинг ухудшается, чтобы график был распределенным.
    streak_penalty = max(0, streak - 3) * 0.12

    return (
        (0.55 * hours_ratio)
        + (0.35 * days_ratio)
        + day_presence_penalty
        + day_load_penalty
        + pace_penalty
        + streak_penalty
        + deterministic_tiebreak
    )


def _pick_employee(
    employees: list[_EmployeeState],
    day: int,
    shift: Shift,
    role_group: str,
    prefer_existing_day: bool,
    day_rank: int,
    total_days: int,
) -> Optional[_EmployeeState]:
    candidates = [emp for emp in employees if _can_take_shift_for_group(emp, day, shift, role_group)]
    if not candidates:
        return None

    scored = []
    for emp in candidates:
        tie = (abs(hash((emp.id, day, shift.key))) % 1000) / 10_000_000
        score = _candidate_score(
            emp,
            day,
            prefer_existing_day,
            tie,
            day_rank=day_rank,
            total_days=total_days,
        )
        scored.append((score, emp))

    scored.sort(key=lambda x: x[0])
    return scored[0][1]


def _pick_from_primary_or_fallback(
    primary_pool: list[_EmployeeState],
    fallback_pool: list[_EmployeeState],
    day: int,
    shift: Shift,
    role_group: str,
    prefer_existing_day: bool,
    day_rank: int,
    total_days: int,
) -> tuple[Optional[_EmployeeState], bool]:
    picked = _pick_employee(
        primary_pool,
        day,
        shift,
        role_group,
        prefer_existing_day,
        day_rank=day_rank,
        total_days=total_days,
    )
    if picked is not None:
        return picked, False

    if fallback_pool:
        picked = _pick_employee(
            fallback_pool,
            day,
            shift,
            role_group,
            prefer_existing_day,
            day_rank=day_rank,
            total_days=total_days,
        )
        if picked is not None:
            return picked, True

    return None, False


def _dedupe_states(states: list[_EmployeeState]) -> list[_EmployeeState]:
    seen: set[tuple[str, str, str]] = set()
    result: list[_EmployeeState] = []
    for emp in states:
        if emp.id in seen:
            continue
        seen.add(emp.id)
        result.append(emp)
    return result


def _assign_shift(emp: _EmployeeState, day: int, shift: Shift, role_group: str) -> None:
    if day not in emp.assigned_days:
        emp.assigned_days.add(day)
        emp.used_days += 1

    shared_same_shift = _can_share_same_shift(
        emp=emp,
        day=day,
        shift_key=shift.key,
        role_group=role_group,
    )
    if not shared_same_shift:
        emp.used_hours += shift.hours
        emp.daily_hours[day] += shift.hours
    emp.daily_shift_keys[day].add(shift.key)
    emp.daily_shift_groups[day][shift.key].add(role_group)


def _build_warnings(
    mandatory_deficits: list[dict[str, Any]],
    optional_missed_count: int,
    cross_restaurant_count: int,
    cross_restaurant_employees: set[str],
) -> list[str]:
    warnings: list[str] = []

    if mandatory_deficits:
        warnings.append(
            f"Обязательные смены с дефицитом: {len(mandatory_deficits)}."
        )

        deficit_df = pd.DataFrame(mandatory_deficits)
        grouped = (
            deficit_df.groupby(["restaurant", "role"], as_index=False)
            .size()
            .sort_values("size", ascending=False)
        )
        for _, row in grouped.head(10).iterrows():
            warnings.append(
                f"Дефицит {int(row['size'])} смен: {row['restaurant']} / {row['role']}."
            )

    if optional_missed_count:
        warnings.append(
            f"Дополнительные (необязательные) слоты без назначения: {optional_missed_count}."
        )

    if cross_restaurant_count:
        warnings.append(
            f"Межресторанные подмены: {cross_restaurant_count}."
        )
        names = sorted(cross_restaurant_employees)
        if names:
            sample = ", ".join(names[:20])
            suffix = " ..." if len(names) > 20 else ""
            warnings.append(f"Сотрудники с межресторанными сменами: {sample}{suffix}")

    return warnings


def generate_schedule(
    employees_df: pd.DataFrame,
    days: list[int],
    weekend_days: Optional[set[int]] = None,
) -> ScheduleResult:
    if employees_df.empty:
        raise ScheduleGenerationError("Нет сотрудников для планирования.")

    if not days:
        raise ScheduleGenerationError("В шаблоне табеля не найдены дни месяца.")

    states = [_EmployeeState(row) for _, row in employees_df.iterrows()]

    by_group: dict[tuple[str, str], list[_EmployeeState]] = defaultdict(list)
    for state in states:
        by_group[(state.restaurant, state.role_group)].append(state)

    restaurants = sorted({state.restaurant for state in states})
    target_groups_set: set[tuple[str, str]] = set(by_group.keys())
    for restaurant in restaurants:
        for group in CORE_MANDATORY_GROUPS:
            target_groups_set.add((restaurant, group))
    group_priority = {
        CASH_ROLE: 0,
        HALL_ROLE: 1,
        KITCHEN_ROLE: 2,
        BAR_ROLE: 3,
    }
    target_groups = sorted(
        target_groups_set,
        key=lambda x: (x[0], group_priority.get(x[1], 99), x[1]),
    )
    kitchen_pool = [state for state in states if state.role_group == KITCHEN_ROLE]
    bar_pool = [state for state in states if state.role_group == BAR_ROLE]
    cash_pool = [state for state in states if state.role_group == CASH_ROLE]
    front_multirole_pool = [state for state in states if _is_front_multirole(state)]

    assignments: list[dict[str, Any]] = []
    mandatory_deficits: list[dict[str, Any]] = []
    optional_missed_count = 0
    cross_restaurant_count = 0
    cross_restaurant_employees: set[str] = set()

    sorted_days = sorted(int(day) for day in days)
    weekend_days_set = set(int(day) for day in (weekend_days or set()))

    total_days = len(sorted_days)

    for day_rank, day in enumerate(sorted_days, start=1):
        for restaurant, role_group in target_groups:
            primary_pool = by_group[(restaurant, role_group)]
            fallback_pool: list[_EmployeeState] = []

            if role_group == KITCHEN_ROLE:
                fallback_pool = [emp for emp in kitchen_pool if emp.restaurant != restaurant]
            elif role_group == BAR_ROLE:
                # Для бара разрешаем:
                # 1) межресторанные барные подмены;
                # 2) подмену универсальным фронт-персоналом (менеджер/администратор),
                #    сначала локально, затем межресторанно.
                fallback_pool = [emp for emp in bar_pool if emp.restaurant != restaurant]
                fallback_pool.extend([emp for emp in front_multirole_pool if emp.restaurant == restaurant])
                fallback_pool.extend([emp for emp in front_multirole_pool if emp.restaurant != restaurant])
                fallback_pool = _dedupe_states(fallback_pool)
            elif role_group == HALL_ROLE:
                # Для зала: при дефиците подменяет универсальный фронт-персонал.
                fallback_pool = [emp for emp in front_multirole_pool if emp.restaurant == restaurant]
                fallback_pool.extend([emp for emp in front_multirole_pool if emp.restaurant != restaurant])
                fallback_pool = _dedupe_states(fallback_pool)
            elif role_group == CASH_ROLE:
                # Для кассы: разрешаем межресторанную подмену кассой
                # и универсальным фронт-персоналом.
                fallback_pool = [emp for emp in cash_pool if emp.restaurant != restaurant]
                fallback_pool.extend([emp for emp in front_multirole_pool if emp.restaurant == restaurant])
                fallback_pool.extend([emp for emp in front_multirole_pool if emp.restaurant != restaurant])
                fallback_pool = _dedupe_states(fallback_pool)

            role_is_mandatory = _is_core_group(role_group) or any(
                _is_mandatory_coverage_role(emp.role_original) for emp in primary_pool
            )

            for shift_key in MANDATORY_SHIFT_KEYS:
                shift = _shift_by_key(shift_key)
                required_slots = _required_slots_for_shift(
                    day=day,
                    shift_key=shift_key,
                    weekend_days=weekend_days_set,
                    role=role_group,
                    is_mandatory_role=role_is_mandatory,
                )
                prefer_existing_day = not role_is_mandatory

                for _slot in range(required_slots):
                    picked, _ = _pick_from_primary_or_fallback(
                        primary_pool=primary_pool,
                        fallback_pool=fallback_pool,
                        day=day,
                        shift=shift,
                        role_group=role_group,
                        prefer_existing_day=prefer_existing_day,
                        day_rank=day_rank,
                        total_days=total_days,
                    )

                    if picked is None:
                        if role_is_mandatory:
                            assignments.append(
                                {
                                    "day": day,
                                    "restaurant": restaurant,
                                    "role": role_group,
                                    "role_group": role_group,
                                    "role_original": "",
                                    "shift": shift.key,
                                    "shift_label": shift.label,
                                    "start": shift.start,
                                    "end": shift.end,
                                    "hours": shift.hours,
                                    "employee": DEFICIT_EMPLOYEE_LABEL,
                                    "employee_home_restaurant": "",
                                    "cross_restaurant": False,
                                    "deficit": True,
                                    "status": "дефицит",
                                }
                            )
                            mandatory_deficits.append(
                                {
                                    "day": day,
                                    "restaurant": restaurant,
                                    "role": role_group,
                                    "shift": shift.label,
                                }
                            )
                        else:
                            optional_missed_count += 1
                        continue

                    _assign_shift(picked, day, shift, role_group)
                    cross_restaurant = picked.restaurant != restaurant
                    if cross_restaurant:
                        cross_restaurant_count += 1
                        cross_restaurant_employees.add(picked.employee)

                    assignments.append(
                        {
                            "day": day,
                            "restaurant": restaurant,
                            "role": picked.role_original,
                            "role_group": role_group,
                            "role_original": picked.role_original,
                            "shift": shift.key,
                            "shift_label": shift.label,
                            "start": shift.start,
                            "end": shift.end,
                            "hours": shift.hours,
                            "employee": picked.employee,
                            "employee_home_restaurant": picked.restaurant,
                            "cross_restaurant": cross_restaurant,
                            "deficit": False,
                            "status": "межресторанная замена" if cross_restaurant else "ok",
                        }
                    )

            for shift_key in OPTIONAL_SHIFT_KEYS:
                shift = _shift_by_key(shift_key)
                picked, _ = _pick_from_primary_or_fallback(
                    primary_pool=primary_pool,
                    fallback_pool=fallback_pool,
                    day=day,
                    shift=shift,
                    role_group=role_group,
                    prefer_existing_day=True,
                    day_rank=day_rank,
                    total_days=total_days,
                )

                if picked is None:
                    optional_missed_count += 1
                    continue

                _assign_shift(picked, day, shift, role_group)
                cross_restaurant = picked.restaurant != restaurant
                if cross_restaurant:
                    cross_restaurant_count += 1
                    cross_restaurant_employees.add(picked.employee)

                assignments.append(
                    {
                            "day": day,
                            "restaurant": restaurant,
                            "role": picked.role_original,
                            "role_group": role_group,
                            "role_original": picked.role_original,
                            "shift": shift.key,
                            "shift_label": shift.label,
                            "start": shift.start,
                            "end": shift.end,
                        "hours": shift.hours,
                        "employee": picked.employee,
                        "employee_home_restaurant": picked.restaurant,
                        "cross_restaurant": cross_restaurant,
                        "deficit": False,
                        "status": "межресторанная замена" if cross_restaurant else "ok",
                    }
                )

    if assignments:
        assignments_df = pd.DataFrame(assignments).sort_values(
            ["day", "restaurant", "role_group", "role", "start", "employee"]
        )
    else:
        assignments_df = pd.DataFrame(
            columns=[
                "day",
                "restaurant",
                "role",
                "role_group",
                "role_original",
                "shift",
                "shift_label",
                "start",
                "end",
                "hours",
                "employee",
                "employee_home_restaurant",
                "cross_restaurant",
                "deficit",
                "status",
            ]
        )

    cross_by_employee: dict[str, int] = {}
    if not assignments_df.empty:
        cross_rows = assignments_df[(~assignments_df["deficit"]) & (assignments_df["cross_restaurant"])]
        cross_by_employee = cross_rows.groupby("employee").size().to_dict()

    summary_rows = []
    for emp in states:
        summary_rows.append(
            {
                "employee": emp.employee,
                "restaurant": emp.restaurant,
                "tab_number": emp.tab_number,
                "role": emp.role_original,
                "role_group": emp.role_group,
                "planned_hours": round(emp.used_hours, 2),
                "planned_days": emp.used_days,
                "max_hours": round(emp.max_hours, 2),
                "max_days": emp.max_days,
                "first_half_pay": round(emp.first_half_pay, 2),
                "second_half_pay": round(emp.second_half_pay, 2),
                "half_preference": emp.half_preference,
                "hours_ok": emp.used_hours <= emp.max_hours + 1e-9,
                "days_ok": emp.used_days <= emp.max_days,
                "cross_restaurant_shifts": int(cross_by_employee.get(emp.employee, 0)),
            }
        )

    summary_df = pd.DataFrame(summary_rows).sort_values(
        ["restaurant", "role", "employee"]
    )

    code_map = {"lunch": "БЛ", "evening": "В", "morning": "У", "day": "Д"}
    code_order = {"У": 0, "Д": 1, "БЛ": 2, "В": 3}

    matrix_index_df = pd.DataFrame(
        [{"employee": emp.employee, "restaurant": emp.restaurant, "role": emp.role_original} for emp in states]
    ).drop_duplicates()

    matrix_base = assignments_df[
        (~assignments_df.get("deficit", False))
        & (assignments_df.get("employee", "") != DEFICIT_EMPLOYEE_LABEL)
    ].copy()

    if matrix_base.empty:
        matrix_df = matrix_index_df.copy()
    else:
        matrix_base["shift_code"] = matrix_base["shift"].map(code_map)
        pivot_df = (
            matrix_base.groupby(["employee", "restaurant", "role", "day"], as_index=False)["shift_code"]
            .agg(lambda values: "+".join(sorted(set(values), key=lambda x: code_order.get(x, 99))))
            .pivot(index=["employee", "restaurant", "role"], columns="day", values="shift_code")
            .reset_index()
        )
        matrix_df = matrix_index_df.merge(pivot_df, on=["employee", "restaurant", "role"], how="left")

    for day in sorted_days:
        if day not in matrix_df.columns:
            matrix_df[day] = ""

    matrix_df = matrix_df.fillna("")
    numeric_columns = [col for col in matrix_df.columns if isinstance(col, int)]
    ordered_cols = ["employee", "restaurant", "role"] + sorted(numeric_columns)
    matrix_df = matrix_df[ordered_cols]

    warnings = _build_warnings(
        mandatory_deficits=mandatory_deficits,
        optional_missed_count=optional_missed_count,
        cross_restaurant_count=cross_restaurant_count,
        cross_restaurant_employees=cross_restaurant_employees,
    )

    return ScheduleResult(
        assignments=assignments_df,
        employee_summary=summary_df,
        matrix=matrix_df,
        warnings=warnings,
    )


def export_schedule_to_excel(
    result: ScheduleResult,
    output_path: Path,
    days: Optional[list[int]] = None,
    weekend_days: Optional[set[int]] = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    matrix_days = sorted({int(day) for day in (days or [])})
    if not matrix_days:
        matrix_days = sorted(
            {
                int(col)
                for col in result.matrix.columns
                if isinstance(col, int) or (isinstance(col, str) and str(col).isdigit())
            }
        )

    # Матрицу строим из той же логики, что и Т-13 (плановые дни/часы из расчетника + приоритет половины месяца).
    preview_df = build_preview_rows_t13_aligned(
        result=result,
        days=matrix_days,
        weekend_days=set(weekend_days or set()),
    )
    if preview_df.empty:
        matrix_download_df = pd.DataFrame(columns=["ФИО сотрудника", "Табельный номер", "Должность"])
    else:
        employee_meta = (
            result.employee_summary[
                ["employee", "tab_number", "role", "restaurant"]
            ]
            .drop_duplicates()
            .sort_values(["employee", "tab_number", "role"], kind="stable")
            .groupby("employee", as_index=False)
            .agg(
                tab_number=("tab_number", "first"),
                role=("role", "first"),
                restaurant=("restaurant", "first"),
            )
        )
        day_hours_by_employee: dict[tuple[str, int], float] = defaultdict(float)
        for _, row in preview_df.iterrows():
            if bool(row.get("deficit", False)):
                continue
            employee = str(row.get("employee", "") or "")
            day = int(pd.to_numeric(row.get("day"), errors="coerce") or 0)
            hours = float(pd.to_numeric(row.get("hours"), errors="coerce") or 0.0)
            if employee and day > 0:
                day_hours_by_employee[(employee, day)] += hours

        matrix_rows: list[dict[str, Any]] = []
        for _, emp_row in employee_meta.iterrows():
            employee = str(emp_row.get("employee", "") or "")
            row: dict[str, Any] = {
                "ФИО сотрудника": employee,
                "Табельный номер": str(emp_row.get("tab_number", "") or ""),
                "Должность": str(emp_row.get("role", "") or ""),
                "Подразделение": str(emp_row.get("restaurant", "") or ""),
            }
            shifts_count = 0
            total_hours = 0.0
            for day in matrix_days:
                day_hours = float(day_hours_by_employee.get((employee, day), 0.0))
                if day_hours > 0:
                    row[f"{day:02d} Явка"] = "Я"
                    row[f"{day:02d} Часы"] = round(day_hours, 2)
                    shifts_count += 1
                    total_hours += day_hours
                else:
                    row[f"{day:02d} Явка"] = "В"
                    row[f"{day:02d} Часы"] = ""
            row["Итого смен"] = shifts_count
            row["Итого часов"] = round(total_hours, 2)
            matrix_rows.append(row)

        matrix_download_df = pd.DataFrame(matrix_rows)
        day_totals_row: dict[str, Any] = {
            "ФИО сотрудника": "ИТОГО ПО ДНЯМ",
            "Табельный номер": "",
            "Должность": "",
            "Подразделение": "",
        }
        total_shifts_overall = 0
        total_hours_overall = 0.0
        for day in matrix_days:
            workers_count = 0
            hours_sum = 0.0
            for _, emp_row in employee_meta.iterrows():
                employee = str(emp_row.get("employee", "") or "")
                day_hours = float(day_hours_by_employee.get((employee, day), 0.0))
                if day_hours > 0:
                    workers_count += 1
                    hours_sum += day_hours
            day_totals_row[f"{day:02d} Явка"] = workers_count if workers_count > 0 else ""
            day_totals_row[f"{day:02d} Часы"] = round(hours_sum, 2) if hours_sum > 0 else ""
            total_shifts_overall += workers_count
            total_hours_overall += hours_sum
        day_totals_row["Итого смен"] = total_shifts_overall if total_shifts_overall > 0 else ""
        day_totals_row["Итого часов"] = round(total_hours_overall, 2) if total_hours_overall > 0 else ""
        matrix_download_df = pd.concat([matrix_download_df, pd.DataFrame([day_totals_row])], ignore_index=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        matrix_download_df.to_excel(writer, index=False, sheet_name="Матрица графика")
        result.assignments.to_excel(writer, index=False, sheet_name="График (сырье)")
        result.employee_summary.to_excel(writer, index=False, sheet_name="Сводка")
        result.matrix.to_excel(writer, index=False, sheet_name="Матрица (тех.)")

        ws_matrix = writer.sheets.get("Матрица графика")
        if ws_matrix is not None and ws_matrix.max_row >= 2:
            right_align = Alignment(horizontal="right", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")
            center_align = Alignment(horizontal="center", vertical="center")

            for col_idx in range(1, ws_matrix.max_column + 1):
                header = str(ws_matrix.cell(row=1, column=col_idx).value or "").strip()
                if header.endswith("Явка"):
                    for row_idx in range(2, ws_matrix.max_row + 1):
                        ws_matrix.cell(row=row_idx, column=col_idx).alignment = right_align
                elif header.endswith("Часы") or header == "Итого часов":
                    for row_idx in range(2, ws_matrix.max_row + 1):
                        ws_matrix.cell(row=row_idx, column=col_idx).alignment = left_align
                elif header == "Итого смен":
                    for row_idx in range(2, ws_matrix.max_row + 1):
                        ws_matrix.cell(row=row_idx, column=col_idx).alignment = center_align


def build_preview_rows_t13_aligned(
    result: ScheduleResult,
    days: list[int],
    weekend_days: Optional[set[int]] = None,
) -> pd.DataFrame:
    sorted_days = sorted({int(day) for day in days})
    weekend_days_set = set(int(day) for day in (weekend_days or set()))

    assignments = result.assignments.copy()
    if assignments.empty:
        assignments = pd.DataFrame(columns=["employee", "day", "restaurant", "deficit", "cross_restaurant"])
    if "deficit" not in assignments.columns:
        assignments["deficit"] = False
    assignments = assignments[(~assignments["deficit"]) & (assignments["employee"] != DEFICIT_EMPLOYEE_LABEL)].copy()
    assignments["day"] = pd.to_numeric(assignments.get("day"), errors="coerce").fillna(0).astype(int)
    assignments["restaurant"] = assignments.get("restaurant", "").astype(str)

    day_restaurants_map: dict[tuple[str, int], set[str]] = defaultdict(set)
    for _, row in assignments.iterrows():
        key = (str(row["employee"]), int(row["day"]))
        day_restaurants_map[key].add(str(row["restaurant"]))

    employee_base = (
        result.employee_summary[
            [
                "employee",
                "restaurant",
                "tab_number",
                "role",
                "role_group",
                "max_days",
                "max_hours",
                "half_preference",
            ]
        ]
        .drop_duplicates()
        .sort_values(["restaurant", "role", "employee"])
        .reset_index(drop=True)
    )

    rows: list[dict[str, Any]] = []
    for _, rec in employee_base.iterrows():
        employee = str(rec["employee"])
        home_restaurant = str(rec["restaurant"])
        role = str(rec["role"])
        role_group = str(rec["role_group"])
        half_preference = str(rec.get("half_preference", "neutral") or "neutral")
        payroll_days_target = int(pd.to_numeric(rec["max_days"], errors="coerce") or 0)
        payroll_hours_target = float(pd.to_numeric(rec["max_hours"], errors="coerce") or 0.0)

        factual_days = sorted({day for (emp_key, day) in day_restaurants_map.keys() if emp_key == employee})
        selected_days = _select_employee_days(
            factual_days=factual_days,
            all_days=sorted_days,
            target_count=payroll_days_target,
            prefer_weekends=_is_core_group(role_group),
            weekend_days=weekend_days_set,
            half_preference=half_preference,
        )

        day_hours = _distribute_hours(payroll_hours_target, len(selected_days))
        hours_map = {day: day_hours[pos] for pos, day in enumerate(selected_days)}

        for day in selected_days:
            worked_restaurants = sorted(day_restaurants_map.get((employee, day), set()))
            factual = bool(worked_restaurants)
            cross = any(rest != home_restaurant for rest in worked_restaurants)
            display_restaurant = worked_restaurants[0] if worked_restaurants else home_restaurant

            status = "ОК"
            if not factual:
                status = "План из расчетного листка"
            elif cross:
                status = "Межресторанная замена"

            rows.append(
                {
                    "day": int(day),
                    "restaurant": display_restaurant,
                    "role": role,
                    "shift_label": "Я",
                    "start": "-",
                    "end": "-",
                    "hours": round(float(hours_map.get(day, 0.0)), 2),
                    "employee": employee,
                    "employee_home_restaurant": home_restaurant,
                    "cross_restaurant": cross,
                    "deficit": False,
                    "status": status,
                }
            )

    # Добавляем отдельные строки дефицита, чтобы их можно было быстро фильтровать в preview.
    deficit_rows = result.assignments[result.assignments.get("deficit", False)].copy()
    if not deficit_rows.empty:
        deficit_rows["day"] = pd.to_numeric(deficit_rows.get("day"), errors="coerce").fillna(0).astype(int)
        for _, row in deficit_rows.iterrows():
            role_value = str(row.get("role", "") or "").strip()
            role_group_value = str(row.get("role_group", "") or "").strip()
            role_display = role_value if role_value else role_group_value
            rows.append(
                {
                    "day": int(row.get("day", 0) or 0),
                    "restaurant": str(row.get("restaurant", "") or ""),
                    "role": role_display,
                    "shift_label": "ДЕФ",
                    "start": "-",
                    "end": "-",
                    "hours": round(float(pd.to_numeric(row.get("hours"), errors="coerce") or 0.0), 2),
                    "employee": DEFICIT_EMPLOYEE_LABEL,
                    "employee_home_restaurant": "-",
                    "cross_restaurant": False,
                    "deficit": True,
                    "status": "ДЕФИЦИТ",
                }
            )

    if not rows:
        return pd.DataFrame(
            columns=[
                "day",
                "restaurant",
                "role",
                "shift_label",
                "start",
                "end",
                "hours",
                "employee",
                "employee_home_restaurant",
                "cross_restaurant",
                "deficit",
                "status",
            ]
        )

    preview_df = pd.DataFrame(rows).sort_values(["day", "restaurant", "role", "employee"]).reset_index(drop=True)
    return preview_df


def _distribute_hours(total_hours: float, day_count: int) -> list[float]:
    if day_count <= 0:
        return []
    if total_hours <= 0:
        return [0.0] * day_count

    units = int(round(total_hours * 2))
    base = units // day_count
    rem = units % day_count

    distributed = [base for _ in range(day_count)]
    for idx in range(rem):
        distributed[idx] += 1

    # Ограничение в 13 часов в день (26 получасов).
    cap = 26
    overflow = 0
    for idx, val in enumerate(distributed):
        if val > cap:
            overflow += val - cap
            distributed[idx] = cap

    idx = 0
    while overflow > 0 and day_count > 0:
        slot = idx % day_count
        if distributed[slot] < cap:
            distributed[slot] += 1
            overflow -= 1
        idx += 1
        if idx > day_count * 100:
            break

    return [val / 2.0 for val in distributed]


def _pick_extra_days(days: list[int], occupied: set[int], count: int) -> list[int]:
    if count <= 0:
        return []

    candidates = [day for day in days if day not in occupied]
    if not candidates:
        return []

    picked: list[int] = []
    step = max(1, len(candidates) // max(1, count))
    idx = 0
    while len(picked) < count and candidates:
        pick = candidates[idx % len(candidates)]
        if pick not in picked:
            picked.append(pick)
        idx += step
        if idx > len(candidates) * 10:
            for day in candidates:
                if len(picked) >= count:
                    break
                if day not in picked:
                    picked.append(day)
            break
    return sorted(picked[:count])


def _pick_evenly_from_days(candidate_days: list[int], count: int) -> list[int]:
    unique_days = sorted(set(int(day) for day in candidate_days))
    if count <= 0 or not unique_days:
        return []
    if count >= len(unique_days):
        return unique_days

    if count == 1:
        return [unique_days[len(unique_days) // 2]]

    n = len(unique_days)
    picked_indices: set[int] = set()
    for i in range(count):
        idx = round(i * (n - 1) / (count - 1))
        picked_indices.add(int(idx))

    # Если из-за округления взяли меньше count, добираем промежуточными индексами.
    cursor = 0
    while len(picked_indices) < count and cursor < n:
        picked_indices.add(cursor)
        cursor += 1

    picked = [unique_days[idx] for idx in sorted(picked_indices)]
    return sorted(picked[:count])


def _select_employee_days(
    factual_days: list[int],
    all_days: list[int],
    target_count: int,
    prefer_weekends: bool,
    weekend_days: set[int],
    half_preference: str = "neutral",
) -> list[int]:
    if target_count <= 0:
        return []

    factual_unique = sorted(set(int(day) for day in factual_days))
    selected = _pick_evenly_from_days(factual_unique, min(target_count, len(factual_unique)))

    if len(selected) < target_count:
        remaining = [day for day in sorted(set(all_days)) if day not in selected]
        if prefer_weekends:
            remaining.sort(key=lambda d: (d not in weekend_days, d))
        extras = _pick_evenly_from_days(remaining, target_count - len(selected))
        selected = sorted(set(selected + extras))

    if len(selected) > target_count:
        selected = _pick_evenly_from_days(selected, target_count)

    # Для core-групп стараемся, чтобы выходных было ощутимо больше, если есть доступные.
    if prefer_weekends and weekend_days and selected:
        weekend_selected = [day for day in selected if day in weekend_days]
        weekend_available = [day for day in all_days if day in weekend_days]
        min_weekend = min(len(weekend_available), max(1, round(target_count * 0.4)))

        if len(weekend_selected) < min_weekend:
            need = min_weekend - len(weekend_selected)
            replace_from = [day for day in selected if day not in weekend_days]
            replace_to = [day for day in weekend_available if day not in selected]
            replace_to = _pick_evenly_from_days(replace_to, need)
            for idx, day in enumerate(replace_to):
                if idx >= len(replace_from):
                    break
                selected.remove(replace_from[idx])
                selected.append(day)
            selected = sorted(set(selected))
            if len(selected) > target_count:
                selected = _pick_evenly_from_days(selected, target_count)

    # Персональный приоритет половины месяца по выплатам:
    # first -> больше дней в 1-15, second -> больше в 16-31.
    preference = str(half_preference or "neutral").strip().lower()
    if preference in {"first", "second"} and selected and target_count > 1:
        first_half_pool = [day for day in all_days if day <= 15]
        second_half_pool = [day for day in all_days if day >= 16]
        selected_first = [day for day in selected if day <= 15]
        selected_second = [day for day in selected if day >= 16]

        target_majority = max(1, round(target_count * 0.6))
        if preference == "first":
            required = min(len(first_half_pool), target_majority)
            if len(selected_first) < required:
                need = required - len(selected_first)
                donor = [day for day in sorted(selected_second)]
                receiver_pool = [day for day in first_half_pool if day not in selected]
                receiver = _pick_evenly_from_days(receiver_pool, need)
                for i, day in enumerate(receiver):
                    if i >= len(donor):
                        break
                    selected.remove(donor[i])
                    selected.append(day)
        else:
            required = min(len(second_half_pool), target_majority)
            if len(selected_second) < required:
                need = required - len(selected_second)
                donor = [day for day in sorted(selected_first)]
                receiver_pool = [day for day in second_half_pool if day not in selected]
                receiver = _pick_evenly_from_days(receiver_pool, need)
                for i, day in enumerate(receiver):
                    if i >= len(donor):
                        break
                    selected.remove(donor[i])
                    selected.append(day)
        selected = sorted(set(selected))
        if len(selected) > target_count:
            selected = _pick_evenly_from_days(selected, target_count)

    return sorted(selected)


def _build_t13_dataframe(result: ScheduleResult, days: list[int]) -> tuple[pd.DataFrame, dict[str, str]]:
    sorted_days = sorted({int(day) for day in days})
    day_columns = [str(day) for day in sorted_days]

    assignments = result.assignments.copy()
    if assignments.empty:
        assignments = pd.DataFrame(
            columns=[
                "employee",
                "day",
                "hours",
                "deficit",
                "cross_restaurant",
                "restaurant",
                "shift",
            ]
        )

    if "day" in assignments.columns:
        assignments["day"] = pd.to_numeric(assignments["day"], errors="coerce").fillna(0).astype(int)
    else:
        assignments["day"] = 0

    if "deficit" not in assignments.columns:
        assignments["deficit"] = False
    else:
        assignments["deficit"] = assignments["deficit"].fillna(False)

    filtered = assignments[
        (~assignments["deficit"]) & (assignments["employee"] != DEFICIT_EMPLOYEE_LABEL)
    ].copy()

    employee_base = (
        result.employee_summary[["employee", "restaurant", "tab_number", "role", "role_group", "half_preference"]]
        .drop_duplicates()
        .sort_values(["restaurant", "role", "employee"])
        .reset_index(drop=True)
    )

    all_restaurants = sorted(
        set(employee_base["restaurant"].astype(str).tolist())
        | set(filtered.get("restaurant", pd.Series(dtype=str)).astype(str).tolist())
    )
    restaurant_codes = {name: f"R{idx:02d}" for idx, name in enumerate(all_restaurants, start=1)}

    day_hours_map: dict[tuple[str, int], float] = {}
    day_details_map: dict[tuple[str, int], list[tuple[str, str]]] = defaultdict(list)

    if not filtered.empty:
        filtered["hours"] = pd.to_numeric(filtered["hours"], errors="coerce").fillna(0.0)
        filtered["restaurant"] = filtered["restaurant"].astype(str)

        shift_to_code = {"lunch": "БЛ", "evening": "В", "morning": "У", "day": "Д"}
        shift_order = {"У": 0, "Д": 1, "БЛ": 2, "В": 3}
        filtered["shift_code"] = filtered["shift"].map(shift_to_code).fillna("Я")

        day_hours_map = (
            filtered.groupby(["employee", "day"], as_index=False)["hours"]
            .sum()
            .set_index(["employee", "day"])["hours"]
            .to_dict()
        )

        grouped = (
            filtered.groupby(["employee", "day", "restaurant"], as_index=False)["shift_code"]
            .agg(lambda vals: "+".join(sorted(set(vals), key=lambda x: shift_order.get(x, 99))))
        )
        for _, row in grouped.iterrows():
            key = (str(row["employee"]), int(row["day"]))
            day_details_map[key].append((str(row["restaurant"]), str(row["shift_code"])))

        for key in day_details_map:
            day_details_map[key] = sorted(day_details_map[key], key=lambda x: x[0])

    rows: list[dict[str, Any]] = []
    total_hours_all = 0.0
    total_worked_days_all = 0
    total_hours_by_day: dict[int, float] = defaultdict(float)
    worked_day_pairs: set[tuple[str, int]] = set()

    for idx, row in employee_base.iterrows():
        employee = str(row["employee"])
        restaurant = str(row["restaurant"])
        role = str(row["role"])
        role_group = str(row.get("role_group", ""))
        half_preference = str(row.get("half_preference", "neutral") or "neutral")
        base_rest_code = restaurant_codes.get(restaurant, "R00")
        summary_row = result.employee_summary[result.employee_summary["employee"] == employee]
        payroll_days_target = 0
        payroll_hours_target = 0.0
        if not summary_row.empty:
            payroll_days_target = int(pd.to_numeric(summary_row.iloc[0]["max_days"], errors="coerce") or 0)
            payroll_hours_target = float(pd.to_numeric(summary_row.iloc[0]["max_hours"], errors="coerce") or 0.0)
            half_preference = str(summary_row.iloc[0].get("half_preference", half_preference) or half_preference)

        row_codes: dict[str, Any] = {
            "№ п/п": idx + 1,
            "Подразделение": f"{restaurant} ({base_rest_code})",
            "Сотрудник": employee,
            "Должность": role,
        }
        row_hours: dict[str, Any] = {
            "№ п/п": "",
            "Подразделение": "",
            "Сотрудник": "",
            "Должность": "часы",
        }

        factual_days = sorted({day for (emp_key, day) in day_details_map.keys() if emp_key == employee})
        selected_days = _select_employee_days(
            factual_days=factual_days,
            all_days=sorted_days,
            target_count=payroll_days_target,
            prefer_weekends=_is_core_group(role_group),
            weekend_days=set(),
            half_preference=half_preference,
        )

        day_hours = _distribute_hours(payroll_hours_target, len(selected_days))
        hours_map = {day: day_hours[idx] for idx, day in enumerate(selected_days)}

        cross_days: list[int] = []

        for day in sorted_days:
            day_col = str(day)
            key = (employee, day)
            hours = float(hours_map.get(day, 0.0))
            details = day_details_map.get(key, [])

            if hours <= 0:
                row_codes[day_col] = ""
                row_hours[day_col] = ""
                continue

            total_hours_all += hours
            total_hours_by_day[day] += hours
            worked_day_pairs.add(key)

            code_parts = []
            day_is_cross = False
            for rest_name, shift_code in details:
                # В унифицированной Т-13 оставляем классический код "Я" в ячейке.
                code_parts.append("Я")
                if rest_name != restaurant:
                    day_is_cross = True

            row_codes[day_col] = "Я"
            row_hours[day_col] = round(hours, 2)
            if day_is_cross:
                cross_days.append(day)

        row_codes["Итого дней"] = payroll_days_target
        row_codes["Итого часов"] = round(payroll_hours_target, 2)
        row_codes["Примечание"] = f"межресторанные дни: {', '.join(str(d) for d in cross_days)}" if cross_days else ""

        row_hours["Итого дней"] = ""
        row_hours["Итого часов"] = round(payroll_hours_target, 2) if payroll_hours_target > 0 else ""
        row_hours["Примечание"] = ""

        rows.append(row_codes)
        rows.append(row_hours)

    total_worked_days_all = len(worked_day_pairs)

    total_row: dict[str, Any] = {
        "№ п/п": "ИТОГО",
        "Подразделение": "",
        "Сотрудник": "",
        "Должность": "",
    }
    for day in sorted_days:
        day_hours = round(total_hours_by_day.get(day, 0.0), 2)
        total_row[str(day)] = day_hours if day_hours > 0 else ""
    total_row["Итого дней"] = total_worked_days_all
    total_row["Итого часов"] = round(total_hours_all, 2)
    total_row["Примечание"] = ""
    rows.append(total_row)

    ordered_columns = (
        ["№ п/п", "Подразделение", "Сотрудник", "Должность"]
        + day_columns
        + ["Итого дней", "Итого часов", "Примечание"]
    )

    df = pd.DataFrame(rows)
    for col in ordered_columns:
        if col not in df.columns:
            df[col] = ""
    df = df[ordered_columns].fillna("")
    return df, restaurant_codes


def _find_t13_day_columns(ws) -> tuple[dict[int, int], dict[int, int]]:
    def _is_hidden_col(col_idx: int) -> bool:
        letter = get_column_letter(col_idx)
        dim = ws.column_dimensions.get(letter)
        return bool(dim.hidden) if dim is not None else False

    def _strict_half_map(row_idx: int, day_from: int, day_to: int) -> Optional[dict[int, int]]:
        day_to_cols: dict[int, list[int]] = defaultdict(list)
        for col in range(1, ws.max_column + 1):
            if _is_hidden_col(col):
                continue
            val = ws.cell(row=row_idx, column=col).value
            day: Optional[int] = None
            if isinstance(val, (int, float)):
                day = int(val)
            elif isinstance(val, str):
                text = val.strip()
                if text.isdigit():
                    day = int(text)
            if day is None:
                continue
            if day_from <= day <= day_to:
                day_to_cols[day].append(col)

        expected_days = list(range(day_from, day_to + 1))
        if not all(day in day_to_cols for day in expected_days):
            return None

        # Выбираем возрастающую по колонкам последовательность для полного диапазона дней.
        chosen: dict[int, int] = {}
        prev_col = -1
        for day in expected_days:
            candidates = sorted(c for c in day_to_cols[day] if c > prev_col)
            if not candidates:
                return None
            chosen_col = candidates[0]
            chosen[day] = chosen_col
            prev_col = chosen_col

        # Отсекаем ложные попадания из правого служебного блока (там обычно большие разрывы).
        cols = [chosen[d] for d in expected_days]
        gaps = [cols[i + 1] - cols[i] for i in range(len(cols) - 1)]
        if not gaps or max(gaps) > 4:
            return None

        return chosen

    best_first: dict[int, int] = {}
    best_second: dict[int, int] = {}
    for row_idx in range(1, min(80, ws.max_row) + 1):
        first_candidate = _strict_half_map(row_idx, 1, 15)
        second_candidate = _strict_half_map(row_idx, 16, 31)
        if first_candidate and (not best_first or min(first_candidate.values()) < min(best_first.values())):
            best_first = first_candidate
        if second_candidate and (not best_second or min(second_candidate.values()) < min(best_second.values())):
            best_second = second_candidate

    if best_first and best_second:
        return best_first, best_second

    # Fallback на старую эвристику (если шаблон нестандартный).
    def _day_map_from_row(row_idx: int) -> dict[int, int]:
        mapped: dict[int, int] = {}
        for col in range(1, ws.max_column + 1):
            if _is_hidden_col(col):
                continue
            val = ws.cell(row=row_idx, column=col).value
            day: Optional[int] = None
            if isinstance(val, (int, float)):
                day = int(val)
            elif isinstance(val, str):
                text = val.strip()
                if text.isdigit():
                    day = int(text)
            if day is not None and 1 <= day <= 31 and day not in mapped:
                mapped[day] = col
        return mapped

    first_map: dict[int, int] = {}
    second_map: dict[int, int] = {}
    for row_idx in range(1, min(60, ws.max_row) + 1):
        mapped = _day_map_from_row(row_idx)
        if len(mapped) < 8:
            continue
        if any(day <= 15 for day in mapped):
            first_map = mapped if len(mapped) > len(first_map) else first_map
        if any(day >= 16 for day in mapped):
            second_map = mapped if len(mapped) > len(second_map) else second_map
    return first_map, second_map


def _find_t13_totals_columns(
    ws,
    first_half_map: dict[int, int],
    second_half_map: dict[int, int],
) -> tuple[Optional[int], Optional[int]]:
    all_day_cols = list(first_half_map.values()) + list(second_half_map.values())
    if not all_day_cols:
        return None, None

    day_grid_end = max(all_day_cols)
    max_scan_col = min(ws.max_column, day_grid_end + 25)
    best_pair: Optional[tuple[int, int]] = None

    for row in range(1, min(140, ws.max_row) + 1):
        found5: list[int] = []
        found6: list[int] = []
        for col in range(day_grid_end + 1, max_scan_col + 1):
            val = ws.cell(row=row, column=col).value
            text = str(val).strip() if val is not None else ""
            if text == "5":
                found5.append(col)
            elif text == "6":
                found6.append(col)
        for c5 in found5:
            for c6 in found6:
                if c6 <= c5:
                    continue
                if c6 - c5 > 6:
                    continue
                pair = (c5, c6)
                if best_pair is None or pair[0] < best_pair[0]:
                    best_pair = pair

    if best_pair is None:
        return None, None
    return best_pair[0], best_pair[1]


def _find_t13_first_employee_row(ws) -> int:
    num_col, fio_col, tab_col, header_row = _find_t13_employee_columns(ws)
    if header_row is not None:
        return header_row + 1

    # 1) Классический случай: уже есть строка сотрудника (цифровой № + ФИО + таб.номер)
    for row in range(1, min(120, ws.max_row) + 1):
        number = str(ws.cell(row=row, column=num_col).value or "").strip()
        fio = str(ws.cell(row=row, column=fio_col).value or "").strip()
        tab_num = str(ws.cell(row=row, column=tab_col).value or "").strip()
        if number.isdigit() and "(" in fio and tab_num:
            return row

    # 2) Шаблонный случай: в строке стоят плейсхолдеры формата {%<...>%}
    for row in range(1, min(180, ws.max_row) + 1):
        c2 = str(ws.cell(row=row, column=num_col).value or "")
        c3 = str(ws.cell(row=row, column=fio_col).value or "")
        c5 = str(ws.cell(row=row, column=tab_col).value or "")
        if "{%" in c2 and "{%" in c3 and "{%" in c5:
            return row
    return 24


def _find_t13_employee_columns(ws) -> tuple[int, int, int, Optional[int]]:
    """
    Возвращает (колонка №п/п, колонка ФИО, колонка таб.номера, строка заголовка с 1/2/3/4).
    Поддерживает разные шаблоны Т-13.
    """
    max_scan_row = min(ws.max_row, 220)
    max_scan_col = min(ws.max_column, 80)
    best: Optional[tuple[int, int, int, int]] = None
    for row in range(1, max_scan_row + 1):
        cols_1 = []
        cols_2 = []
        cols_3 = []
        cols_4 = []
        for col in range(1, max_scan_col + 1):
            txt = str(ws.cell(row=row, column=col).value or "").strip()
            if txt == "1":
                cols_1.append(col)
            elif txt == "2":
                cols_2.append(col)
            elif txt == "3":
                cols_3.append(col)
            elif txt == "4":
                cols_4.append(col)
        for c1 in cols_1:
            for c2 in cols_2:
                if c2 <= c1:
                    continue
                for c3 in cols_3:
                    if c3 <= c2:
                        continue
                    for c4 in cols_4:
                        if c4 <= c3:
                            continue
                        # Отсеиваем случайные строки с большими разрывами.
                        if c4 - c1 > 16:
                            continue
                        cand = (c1, c2, c3, row)
                        if best is None:
                            best = cand
                        else:
                            # Предпочитаем более "левый" заголовок.
                            if cand[0] < best[0] or (cand[0] == best[0] and cand[3] < best[3]):
                                best = cand
    if best is not None:
        return best[0], best[1], best[2], best[3]
    # Дефолт для старого шаблона.
    return 2, 3, 5, None


def _find_t13_footer_row(ws, start_scan_row: int) -> Optional[int]:
    """
    Ищет начало служебной части бланка Т-13 (подписи/реквизиты организации),
    ниже которой нельзя размещать сотрудников.
    """
    markers = (
        "ответствен",
        "наименование организац",
        "утверждена",
        "постановлен",
        "общество с ограниченной ответственностью",
    )

    max_scan_row = ws.max_row
    max_scan_col = min(ws.max_column, 30)
    for row in range(max(start_scan_row, 1), max_scan_row + 1):
        row_text_parts: list[str] = []
        for col in range(1, max_scan_col + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            text = str(value).strip().lower()
            text = re.sub(r"\s+", " ", text)
            if text:
                row_text_parts.append(text)
        if not row_text_parts:
            continue
        row_text = " ".join(row_text_parts)
        if any(marker in row_text for marker in markers):
            return row
    return None


def _resolve_cell_anchor(ws, row: int, col: int) -> tuple[int, int]:
    cell = ws.cell(row=row, column=col)
    if cell.__class__.__name__ != "MergedCell":
        return row, col
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return merged.min_row, merged.min_col
    return row, col


def _find_merged_range_for_cell(ws, row: int, col: int):
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return merged
    return None


def _unmerge_with_style_copy(ws, merged_range) -> None:
    from copy import copy

    min_row, min_col, max_row, max_col = (
        merged_range.min_row,
        merged_range.min_col,
        merged_range.max_row,
        merged_range.max_col,
    )
    anchor = ws.cell(row=min_row, column=min_col)
    anchor_style = copy(anchor._style)
    anchor_font = copy(anchor.font)
    anchor_fill = copy(anchor.fill)
    anchor_border = copy(anchor.border)
    anchor_alignment = copy(anchor.alignment)
    anchor_number_format = anchor.number_format
    anchor_protection = copy(anchor.protection)

    ws.unmerge_cells(str(merged_range))
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell._style = copy(anchor_style)
            cell.font = copy(anchor_font)
            cell.fill = copy(anchor_fill)
            cell.border = copy(anchor_border)
            cell.alignment = copy(anchor_alignment)
            cell.number_format = anchor_number_format
            cell.protection = copy(anchor_protection)
            if not (r == min_row and c == min_col):
                cell.value = None
                cell.comment = None


def _normalize_day_grid_merges_for_block(ws, block_start: int, first_half_map: dict[int, int], second_half_map: dict[int, int]) -> None:
    """
    В некоторых шаблонах встречаются merge-диапазоны, перекрывающие
    пары строк кода/часов. Это приводит к пропуску часов. Нормализуем
    такие merge внутри сетки дней текущего блока сотрудника.
    """
    # Нормализуем только проблемные merge (когда код и часы попадают в одну anchor-ячейку),
    # не трогая остальную структуру шаблона.

    for _day, col in first_half_map.items():
        a_code = _resolve_cell_anchor(ws, block_start, col)
        a_hours = _resolve_cell_anchor(ws, block_start + 1, col)
        if a_code == a_hours:
            mr = _find_merged_range_for_cell(ws, block_start + 1, col) or _find_merged_range_for_cell(ws, block_start, col)
            if mr is not None:
                _unmerge_with_style_copy(ws, mr)

    for _day, col in second_half_map.items():
        a_code = _resolve_cell_anchor(ws, block_start + 2, col)
        a_hours = _resolve_cell_anchor(ws, block_start + 3, col)
        if a_code == a_hours:
            mr = _find_merged_range_for_cell(ws, block_start + 3, col) or _find_merged_range_for_cell(ws, block_start + 2, col)
            if mr is not None:
                _unmerge_with_style_copy(ws, mr)


def _set_cell_value_safe(ws, row: int, col: int, value: Any) -> None:
    r, c = _resolve_cell_anchor(ws, row, col)
    ws.cell(row=r, column=c).value = value


def _set_cell_comment_safe(ws, row: int, col: int, comment: Optional[Comment]) -> None:
    r, c = _resolve_cell_anchor(ws, row, col)
    ws.cell(row=r, column=c).comment = comment


def _set_cell_alignment_safe(ws, row: int, col: int, alignment: Alignment) -> None:
    r, c = _resolve_cell_anchor(ws, row, col)
    ws.cell(row=r, column=c).alignment = alignment


def _ensure_continuous_table_for_general_sheet(ws, start_row: int, required_blocks: int) -> list[int]:
    """
    Для листа 'Т-13 общий' формирует непрерывную табличную часть (без повторных шапок страниц):
    при нехватке строк добавляет блоки по 4 строки перед первым футером.
    """
    footer_row = _find_t13_footer_row(ws, start_row + 1)
    if footer_row is None:
        return [start_row + idx * 4 for idx in range(required_blocks)]

    available_blocks = max(0, (footer_row - start_row) // 4)
    if required_blocks <= available_blocks:
        return [start_row + idx * 4 for idx in range(required_blocks)]

    missing_blocks = required_blocks - available_blocks
    rows_to_insert = missing_blocks * 4

    # Паттерн стилей берём из первого блока сотрудника.
    style_source_rows = [start_row + i for i in range(4)]
    max_col = ws.max_column

    # Запоминаем merge-паттерны для одного блока.
    block_merges: list[tuple[int, int, int, int]] = []
    for merged in ws.merged_cells.ranges:
        if start_row <= merged.min_row <= merged.max_row <= start_row + 3:
            block_merges.append(
                (
                    merged.min_row - start_row,
                    merged.max_row - start_row,
                    merged.min_col,
                    merged.max_col,
                )
            )

    ws.insert_rows(footer_row, amount=rows_to_insert)

    # Копируем стиль/формат первого блока на новые строки.
    from copy import copy

    for i in range(rows_to_insert):
        target_row = footer_row + i
        src_row = style_source_rows[i % 4]
        if ws.row_dimensions.get(src_row) and ws.row_dimensions[src_row].height is not None:
            ws.row_dimensions[target_row].height = ws.row_dimensions[src_row].height
        for col in range(1, max_col + 1):
            src_cell = ws.cell(row=src_row, column=col)
            dst_cell = ws.cell(row=target_row, column=col)
            dst_cell._style = copy(src_cell._style)
            dst_cell.number_format = src_cell.number_format
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.border = copy(src_cell.border)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.protection = copy(src_cell.protection)
            dst_cell.value = None
            dst_cell.comment = None

    # Полностью пересобираем merge-структуру в рабочей зоне сотрудников,
    # чтобы не оставались "чужие" объединения от следующих страниц шаблона.
    area_end = start_row + required_blocks * 4 - 1
    to_unmerge = []
    for merged in ws.merged_cells.ranges:
        if not (merged.max_row < start_row or merged.min_row > area_end):
            to_unmerge.append(str(merged))
    for ref in to_unmerge:
        ws.unmerge_cells(ref)

    # Восстанавливаем merge-структуру блоками по 4 строки в нужной зоне.
    for block_idx in range(required_blocks):
        base_row = start_row + block_idx * 4
        for min_off, max_off, min_col, max_col in block_merges:
            ws.merge_cells(
                start_row=base_row + min_off,
                end_row=base_row + max_off,
                start_column=min_col,
                end_column=max_col,
            )

    return [start_row + idx * 4 for idx in range(required_blocks)]


def _set_t13_report_period(ws, year: Optional[int], month: Optional[int]) -> None:
    if not year or not month:
        return
    from_text = f"01.{int(month):02d}.{int(year)}"
    last_day = calendar.monthrange(int(year), int(month))[1]
    to_text = f"{int(last_day):02d}.{int(month):02d}.{int(year)}"

    max_scan_row = min(ws.max_row, 180)
    max_scan_col = min(ws.max_column, 260)
    c_col: Optional[int] = None
    po_col: Optional[int] = None
    target_row: Optional[int] = None

    for row in range(1, max_scan_row + 1):
        row_tokens = {}
        for col in range(1, max_scan_col + 1):
            text = str(ws.cell(row=row, column=col).value or "").strip().lower()
            if text:
                row_tokens[col] = text
        if not row_tokens:
            continue
        for col, text in row_tokens.items():
            if text == "с":
                c_col = col
            elif text == "по":
                po_col = col
        if c_col is not None and po_col is not None:
            target_row = row + 1
            break

    if target_row is not None and c_col is not None and po_col is not None:
        _set_cell_value_safe(ws, target_row, c_col, from_text)
        _set_cell_value_safe(ws, target_row, po_col, to_text)


def _find_t13_block_rows(
    ws,
    start_row: int,
    first_half_map: dict[int, int],
    second_half_map: dict[int, int],
    num_col: int = 2,
    fio_col: int = 3,
    tab_col: int = 5,
) -> list[int]:
    def _writable(row: int, col: int) -> bool:
        return ws.cell(row=row, column=col).__class__.__name__ != "MergedCell"

    def _block_has_valid_day_grid(block_start: int) -> bool:
        # Для шаблонов с merge-ячейками не требуем writable в каждой ячейке дня:
        # запись выполняется через _set_cell_value_safe в anchor merged-range.
        if block_start + 3 > ws.max_row:
            return False
        if not first_half_map or not second_half_map:
            return False
        return True

    # 1) Основной путь: берем "якорные" строки сотрудников по шаблонным номерам/ФИО.
    anchor_rows: list[int] = []
    for row in range(max(1, start_row - 4), ws.max_row - 3):
        number = str(ws.cell(row=row, column=num_col).value or "").strip()
        fio = str(ws.cell(row=row, column=fio_col).value or "").strip()
        tab_num = str(ws.cell(row=row, column=tab_col).value or "").strip()
        if not number.isdigit():
            continue
        if not fio or fio.isdigit():
            continue
        if not tab_num:
            continue
        if not (_writable(row, num_col) and _writable(row, fio_col) and _writable(row, tab_col)):
            continue
        anchor_rows.append(row)
    if anchor_rows:
        return sorted(anchor_rows)

    # 2) Fallback: сканируем по всем строкам, чтобы ловить шаблоны,
    # где первый блок начинается не на start_row, а на +1/+2.
    rows: list[int] = []
    row = max(1, start_row - 4)
    while row <= ws.max_row - 3:
        c2 = ws.cell(row=row, column=num_col).value
        c3 = ws.cell(row=row, column=fio_col).value
        c5 = ws.cell(row=row, column=tab_col).value
        c2_txt = str(c2 or "").strip()
        c3_txt = str(c3 or "").strip()
        c5_txt = str(c5 or "").strip()
        has_left_part = bool(c2_txt) and bool(c3_txt) and bool(c5_txt) and not c3_txt.isdigit()
        if has_left_part and _writable(row, num_col) and _writable(row, fio_col) and _writable(row, tab_col) and _block_has_valid_day_grid(row):
            if not rows or row - rows[-1] >= 4:
                rows.append(row)
        row += 1
    return rows


def _fill_t13_template_sheet(
    ws,
    result: ScheduleResult,
    days: list[int],
    restaurant_codes: dict[str, str],
    weekend_days: Optional[set[int]] = None,
    sort_mode: str = "by_restaurant",
    continuous_table: bool = False,
    filter_restaurant: Optional[str] = None,
) -> None:
    sorted_days = sorted({int(day) for day in days})
    weekend_days_set = set(int(day) for day in (weekend_days or set()))
    first_half_map, second_half_map = _find_t13_day_columns(ws)
    totals_col_5, totals_col_6 = _find_t13_totals_columns(ws, first_half_map, second_half_map)
    num_col, fio_col, tab_col, _header_row = _find_t13_employee_columns(ws)
    start_row = _find_t13_first_employee_row(ws)
    def _writable(row: int, col: int) -> bool:
        return ws.cell(row=row, column=col).__class__.__name__ != "MergedCell"

    pre_employee_base = result.employee_summary[
        [
            "employee",
            "restaurant",
            "tab_number",
            "role",
            "role_group",
            "max_days",
            "max_hours",
            "half_preference",
        ]
    ].drop_duplicates()
    if filter_restaurant:
        pre_employee_base = pre_employee_base[
            pre_employee_base["restaurant"].astype(str) == str(filter_restaurant)
        ].reset_index(drop=True)
    needed_blocks = len(pre_employee_base)

    block_rows = _find_t13_block_rows(
        ws=ws,
        start_row=start_row,
        first_half_map=first_half_map,
        second_half_map=second_half_map,
        num_col=num_col,
        fio_col=fio_col,
        tab_col=tab_col,
    )

    preview_df = build_preview_rows_t13_aligned(
        result=result,
        days=sorted_days,
        weekend_days=weekend_days_set,
    )
    preview_df = preview_df[(~preview_df.get("deficit", False)) & (preview_df.get("employee", "") != DEFICIT_EMPLOYEE_LABEL)].copy()
    preview_df["day"] = pd.to_numeric(preview_df.get("day"), errors="coerce").fillna(0).astype(int)
    preview_df["restaurant"] = preview_df.get("restaurant", "").astype(str)

    details_map: dict[tuple[str, int], list[str]] = defaultdict(list)
    for _, row in preview_df.iterrows():
        key = (str(row["employee"]), int(row["day"]))
        details_map[key].append(str(row["restaurant"]))
    day_hours_map: dict[tuple[str, int], float] = {}
    if not preview_df.empty and "hours" in preview_df.columns:
        day_hours_map = (
            preview_df.assign(hours=pd.to_numeric(preview_df["hours"], errors="coerce").fillna(0.0))
            .groupby(["employee", "day"], as_index=False)["hours"]
            .sum()
            .set_index(["employee", "day"])["hours"]
            .to_dict()
        )

    employee_base = pre_employee_base.copy()
    if sort_mode == "alphabetical":
        employee_base = employee_base.sort_values(["employee", "restaurant", "role"]).reset_index(drop=True)
    else:
        employee_base = employee_base.sort_values(["restaurant", "role", "employee"]).reset_index(drop=True)

    if not block_rows:
        # Шаблон без предзаполненных сотрудников: стартуем с первой найденной строки блока.
        block_rows = [start_row]

    # Для общего листа всегда строим непрерывную таблицу сотрудников без
    # промежуточных шапок/подвалов страниц шаблона.
    if continuous_table and block_rows:
        expanded_rows = _ensure_continuous_table_for_general_sheet(
            ws=ws,
            start_row=block_rows[0],
            required_blocks=len(employee_base),
        )
        if expanded_rows:
            block_rows = expanded_rows
    # Если шаблон содержит мало строк-блоков (например только один плейсхолдер),
    # автоматически наращиваем непрерывную табличную часть.
    elif len(block_rows) < len(employee_base) and block_rows:
        expanded_rows = _ensure_continuous_table_for_general_sheet(
            ws=ws,
            start_row=block_rows[0],
            required_blocks=len(employee_base),
        )
        if expanded_rows:
            block_rows = expanded_rows

    max_blocks = len(block_rows)
    fill_count = min(len(employee_base), max_blocks)

    for r in block_rows:
        # Левая часть строки сотрудника.
        for col in [num_col, fio_col]:
            _set_cell_value_safe(ws, r, col, None)
        _set_cell_value_safe(ws, r, tab_col, None)
        mapped_cols = list({**first_half_map, **second_half_map}.values())
        if mapped_cols:
            day_grid_start = max(1, min(mapped_cols))
            day_grid_end = min(ws.max_column, max(mapped_cols))
        else:
            day_grid_start, day_grid_end = 9, ws.max_column

        for row_idx in (r, r + 1, r + 2, r + 3):
            # Очищаем всю правую часть блока сотрудника, чтобы удалить "зашитые"
            # числа шаблона (например 136/96/40/12), которые могут вводить в заблуждение.
            for col in range(day_grid_start, ws.max_column + 1):
                _set_cell_value_safe(ws, row_idx, col, None)
                _set_cell_comment_safe(ws, row_idx, col, None)

    for idx in range(fill_count):
        rec = employee_base.iloc[idx]
        employee = str(rec["employee"])
        restaurant = str(rec["restaurant"])
        tab_number = str(rec.get("tab_number", "") or "")
        role = str(rec["role"])
        role_group = str(rec["role_group"])
        half_preference = str(rec.get("half_preference", "neutral") or "neutral")
        payroll_days_target = int(pd.to_numeric(rec["max_days"], errors="coerce") or 0)
        payroll_hours_target = float(pd.to_numeric(rec["max_hours"], errors="coerce") or 0.0)

        r = block_rows[idx]
        _set_cell_value_safe(ws, r, num_col, str(idx + 1))
        _set_cell_value_safe(ws, r, fio_col, f"{employee}\n({role})")
        _set_cell_value_safe(ws, r, tab_col, tab_number if tab_number else "")
        assigned_days = sorted({day for (emp_key, day) in day_hours_map.keys() if emp_key == employee})
        if assigned_days:
            selected_days = assigned_days
            hours_map = {
                day: round(float(day_hours_map.get((employee, day), 0.0)), 2)
                for day in selected_days
            }
        else:
            factual_days = sorted({day for (emp_key, day) in details_map.keys() if emp_key == employee})
            selected_days = _select_employee_days(
                factual_days=factual_days,
                all_days=sorted_days,
                target_count=payroll_days_target,
                prefer_weekends=_is_core_group(role_group),
                weekend_days=weekend_days_set,
                half_preference=half_preference,
            )
            day_hours = _distribute_hours(payroll_hours_target, len(selected_days))
            hours_map = {day: day_hours[pos] for pos, day in enumerate(selected_days)}
        first_half_days = [day for day in selected_days if day <= 15]
        second_half_days = [day for day in selected_days if day >= 16]
        first_half_days_count = len(first_half_days)
        second_half_days_count = len(second_half_days)
        first_half_hours = round(sum(float(hours_map.get(day, 0.0)) for day in first_half_days), 2)
        second_half_hours = round(sum(float(hours_map.get(day, 0.0)) for day in second_half_days), 2)
        total_days = first_half_days_count + second_half_days_count
        total_hours = round(first_half_hours + second_half_hours, 2)

        if totals_col_5 is not None:
            if _writable(r, totals_col_5):
                _set_cell_value_safe(ws, r, totals_col_5, first_half_days_count if first_half_days_count > 0 else "")
            if _writable(r + 1, totals_col_5):
                _set_cell_value_safe(ws, r + 1, totals_col_5, first_half_hours if first_half_hours > 0 else "")
            if _writable(r + 2, totals_col_5):
                _set_cell_value_safe(
                    ws,
                    r + 2,
                    totals_col_5,
                    second_half_days_count if second_half_days_count > 0 else "",
                )
            if _writable(r + 3, totals_col_5):
                _set_cell_value_safe(ws, r + 3, totals_col_5, second_half_hours if second_half_hours > 0 else "")

        if totals_col_6 is not None:
            if _writable(r, totals_col_6):
                _set_cell_value_safe(ws, r, totals_col_6, total_days if total_days > 0 else "")
            if _writable(r + 1, totals_col_6):
                _set_cell_value_safe(ws, r + 1, totals_col_6, "")
            if _writable(r + 2, totals_col_6):
                _set_cell_value_safe(ws, r + 2, totals_col_6, total_hours if total_hours > 0 else "")
            if _writable(r + 3, totals_col_6):
                _set_cell_value_safe(ws, r + 3, totals_col_6, "")

        for day in sorted_days:
            if day <= 15:
                col = first_half_map.get(day)
                code_row, hours_row = r, r + 1
            else:
                col = second_half_map.get(day)
                code_row, hours_row = r + 2, r + 3
            if col is None:
                continue
            code_col = col
            hours_col = col

            if day not in hours_map:
                if code_col is not None:
                    _set_cell_value_safe(ws, code_row, code_col, "В")
                    _set_cell_comment_safe(ws, code_row, code_col, None)
                if hours_col is not None:
                    _set_cell_value_safe(ws, hours_row, hours_col, None)
                continue

            key = (employee, day)
            worked_restaurants = sorted(set(details_map.get(key, [])))
            if code_col is not None:
                if worked_restaurants:
                    _set_cell_value_safe(ws, code_row, code_col, "Я")
                    if any(rest != restaurant for rest in worked_restaurants):
                        note = "Межресторанная подмена: " + ", ".join(worked_restaurants)
                        _set_cell_comment_safe(ws, code_row, code_col, Comment(note, "tabel"))
                    else:
                        _set_cell_comment_safe(ws, code_row, code_col, None)
                else:
                    _set_cell_value_safe(ws, code_row, code_col, "Я")
                    _set_cell_comment_safe(ws, code_row, code_col, None)
            if hours_col is not None:
                _set_cell_value_safe(ws, hours_row, hours_col, round(float(hours_map[day]), 2))

    ws.cell(row=2, column=1, value=f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}")

    if continuous_table:
        # На непрерывном листе явно убираем скрытие строк, оставшееся от шаблона.
        for row_idx in range(1, ws.max_row + 1):
            if ws.row_dimensions[row_idx].hidden:
                ws.row_dimensions[row_idx].hidden = False


def _clear_structural_subdivision_header(ws) -> None:
    max_scan_row = min(60, ws.max_row)
    max_scan_col = min(40, ws.max_column)
    for row in range(1, max_scan_row + 1):
        for col in range(1, max_scan_col + 1):
            value = ws.cell(row=row, column=col).value
            text = str(value or "").strip().lower()
            if "структурное подразделение" in text:
                ws.cell(row=row, column=col, value="")


def _clear_structural_subdivision_values(ws, restaurant_names: list[str]) -> None:
    normalized_restaurants = {
        re.sub(r"\s+", " ", str(name or "").strip().lower())
        for name in restaurant_names
        if str(name or "").strip()
    }
    if not normalized_restaurants:
        return

    max_scan_row = min(220, ws.max_row)
    max_scan_col = min(80, ws.max_column)
    for row in range(1, max_scan_row + 1):
        for col in range(1, max_scan_col + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            text = re.sub(r"\s+", " ", str(value).strip().lower())
            if text in normalized_restaurants:
                ws.cell(row=row, column=col, value="")


def _clear_comments_in_merged_non_anchor_cells(ws) -> None:
    for merged in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged.min_row, merged.min_col, merged.max_row, merged.max_col
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if row == min_row and col == min_col:
                    continue
                cell = ws.cell(row=row, column=col)
                if cell.comment is not None:
                    cell.comment = None


def _remove_common_sheet_gap_rows_by_context(ws) -> None:
    """
    Сжимает общую таблицу по "якорям" сотрудников:
    между соседними сотрудниками оставляет только 4 строки блока,
    удаляя все промежуточные служебные шапки/подвалы страниц.
    """

    def _employee_anchor_rows() -> list[int]:
        anchors: list[int] = []
        for row in range(1, ws.max_row + 1):
            number = str(ws.cell(row=row, column=2).value or "").strip()
            fio = str(ws.cell(row=row, column=3).value or "").strip()
            tab_num = str(ws.cell(row=row, column=5).value or "").strip()
            if number.isdigit() and "(" in fio and tab_num.isdigit():
                anchors.append(row)
        return anchors

    anchors = _employee_anchor_rows()
    if len(anchors) < 2:
        return

    # Скрываем (а не удаляем) служебные разрывы страниц между соседними сотрудниками.
    # Это сохраняет merge-структуру шаблона и не ломает сетку.
    for cur_anchor, next_anchor in zip(anchors, anchors[1:]):
        gap = next_anchor - cur_anchor
        if gap <= 4:
            continue
        hide_start = cur_anchor + 4
        hide_end = next_anchor - 1
        for row in range(hide_start, hide_end + 1):
            ws.row_dimensions[row].hidden = True


def export_t13_to_excel(
    result: ScheduleResult,
    days: list[int],
    output_path: Path,
    template_bytes: Optional[bytes] = None,
    weekend_days: Optional[set[int]] = None,
    period_year: Optional[int] = None,
    period_month: Optional[int] = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if template_bytes:
        from openpyxl import load_workbook

        template_wb = load_workbook(BytesIO(template_bytes))
        ws = template_wb[template_wb.sheetnames[0]]
        # Общий лист создаем из чистого шаблона (до заполнения основного),
        # чтобы не наследовать уже проставленные значения/комментарии.
        ws_all = template_wb.copy_worksheet(ws)
        ws_all.title = "Т-13 общий"

        all_restaurants = sorted(result.employee_summary["restaurant"].astype(str).unique().tolist())
        restaurant_codes = {name: f"R{idx:02d}" for idx, name in enumerate(all_restaurants, start=1)}
        _fill_t13_template_sheet(
            ws=ws,
            result=result,
            days=days,
            restaurant_codes=restaurant_codes,
            weekend_days=weekend_days,
            sort_mode="by_restaurant",
            continuous_table=False,
        )

        # Листы по каждому ресторану (внутри — только сотрудники подразделения).
        for restaurant_name in all_restaurants:
            ws_rest = template_wb.copy_worksheet(ws_all)
            code = restaurant_codes.get(restaurant_name, "R00")
            ws_rest.title = f"{code}"
            _fill_t13_template_sheet(
                ws=ws_rest,
                result=result,
                days=days,
                restaurant_codes=restaurant_codes,
                weekend_days=weekend_days,
                sort_mode="by_restaurant",
                continuous_table=False,
                filter_restaurant=restaurant_name,
            )
            _set_t13_report_period(ws_rest, period_year, period_month)
            _clear_comments_in_merged_non_anchor_cells(ws_rest)

        # Дополнительный лист: общий табель по юрлицу, сплошным списком по алфавиту.
        _fill_t13_template_sheet(
            ws=ws_all,
            result=result,
            days=days,
            restaurant_codes=restaurant_codes,
            weekend_days=weekend_days,
            sort_mode="alphabetical",
            continuous_table=True,
        )
        _clear_structural_subdivision_header(ws_all)
        _clear_structural_subdivision_values(ws_all, all_restaurants)
        _set_t13_report_period(ws, period_year, period_month)
        _set_t13_report_period(ws_all, period_year, period_month)
        _clear_comments_in_merged_non_anchor_cells(ws)
        _clear_comments_in_merged_non_anchor_cells(ws_all)
        # Базовый лист шаблона (например "стр1") не нужен в финальной выгрузке:
        # оставляем листы по ресторанам + общий.
        if ws in template_wb.worksheets:
            template_wb.remove(ws)
        template_wb.save(output_path)
        return

    t13_df, restaurant_codes = _build_t13_dataframe(result=result, days=days)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        sheet_name = "Т-13"
        start_row = 4
        t13_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=start_row)

        ws = writer.sheets[sheet_name]
        max_col = ws.max_column
        max_row = ws.max_row

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.cell(
            row=1,
            column=1,
            value="Табель учета рабочего времени (Унифицированная форма № Т-13)",
        )
        ws.cell(row=2, column=1, value=f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
        if restaurant_codes:
            legend = "; ".join(f"{code}={name}" for name, code in restaurant_codes.items())
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)
            ws.cell(row=3, column=1, value=f"Коды подразделений: {legend}")

        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=2, column=1).font = Font(italic=True, size=10)
        ws.cell(row=3, column=1).font = Font(size=9, color="404040")
        ws.cell(row=3, column=1).alignment = Alignment(horizontal="left")

        header_row = start_row + 1
        header_fill = PatternFill(start_color="DCEBFA", end_color="DCEBFA", fill_type="solid")
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = header_fill

        ws.freeze_panes = "E6"

        width_map = {
            1: 8,   # №
            2: 28,  # Подразделение
            3: 30,  # Сотрудник
            4: 24,  # Должность
        }
        for col_idx in range(1, max_col + 1):
            width = width_map.get(col_idx, 11)
            if col_idx > max_col - 3:
                width = 14
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        cross_fill = PatternFill(start_color="FFF7DC", end_color="FFF7DC", fill_type="solid")

        for row_idx in range(header_row + 1, max_row + 1):
            role_value = str(ws.cell(row=row_idx, column=4).value or "").strip().lower()
            if role_value == "часы":
                for col_idx in range(1, max_col + 1):
                    ws.cell(row=row_idx, column=col_idx).font = Font(size=9, color="404040")
                    ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                continue

            if str(ws.cell(row=row_idx, column=1).value).strip().upper() == "ИТОГО":
                for col_idx in range(1, max_col + 1):
                    ws.cell(row=row_idx, column=col_idx).font = Font(bold=True)
                    ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                continue

            base_rest = str(ws.cell(row=row_idx, column=2).value or "")
            base_match = re.search(r"\((R\d{2})\)\s*$", base_rest)
            base_code = base_match.group(1) if base_match else ""
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if col_idx <= 4:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            first_day_col = 5
            last_day_col = max_col - 3
            for col_idx in range(first_day_col, last_day_col + 1):
                code_cell = ws.cell(row=row_idx, column=col_idx)
                code_text = str(code_cell.value or "").strip()
                if not code_text:
                    continue

                day_codes = re.findall(r"@([A-Z]\d{2})", code_text)
                if day_codes and any(code != base_code for code in day_codes):
                    code_cell.fill = cross_fill
                    next_role = str(ws.cell(row=row_idx + 1, column=4).value or "").strip().lower()
                    if next_role == "часы":
                        ws.cell(row=row_idx + 1, column=col_idx).fill = cross_fill


def export_t13_to_pdf(
    result: ScheduleResult,
    days: list[int],
    output_path: Path,
    period_year: Optional[int] = None,
    period_month: Optional[int] = None,
) -> None:
    """
    Экспорт заполненного Т-13 в PDF (табличное представление).
    Структура данных синхронизирована с Excel-экспортом через _build_t13_dataframe.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    output_path.parent.mkdir(parents=True, exist_ok=True)
    t13_df, _restaurant_codes = _build_t13_dataframe(result=result, days=days)

    # Подбираем шрифт с кириллицей (иначе в PDF будут "квадраты").
    font_name = "Helvetica"
    font_candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/Library/Fonts/Arial Unicode.ttf",
    ]
    for font_path in font_candidates:
        p = Path(font_path)
        if p.exists():
            try:
                pdfmetrics.registerFont(TTFont("T13Sans", str(p)))
                font_name = "T13Sans"
                break
            except Exception:
                continue

    sorted_days = sorted({int(day) for day in days})
    day_cols = [str(day) for day in sorted_days]
    columns = ["№ п/п", "Подразделение", "Сотрудник", "Должность", *day_cols, "Итого дней", "Итого часов"]

    # Готовим данные таблицы (без служебной колонки "Примечание").
    safe_df = t13_df.copy()
    for col in columns:
        if col not in safe_df.columns:
            safe_df[col] = ""
    safe_df = safe_df[columns].fillna("")

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleSmall",
        parent=styles["Heading2"],
        fontName=font_name,
        fontSize=12,
        leading=14,
        alignment=1,
    )
    meta_style = ParagraphStyle(
        "Meta",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=9,
        leading=11,
    )

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(A3),
        leftMargin=6 * mm,
        rightMargin=6 * mm,
        topMargin=6 * mm,
        bottomMargin=6 * mm,
        title="Табель Т-13",
    )

    period_text = ""
    if period_year and period_month:
        last_day = calendar.monthrange(int(period_year), int(period_month))[1]
        period_text = f"Период: 01.{int(period_month):02d}.{int(period_year)} - {last_day:02d}.{int(period_month):02d}.{int(period_year)}"

    story = [
        Paragraph("Табель учета рабочего времени (Унифицированная форма Т-13)", title_style),
        Spacer(1, 2 * mm),
        Paragraph(period_text if period_text else "Период: из расчетного листа", meta_style),
        Spacer(1, 1 * mm),
        Paragraph(f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}", meta_style),
        Spacer(1, 2 * mm),
    ]

    header = columns
    body_rows = [header]
    for _, row in safe_df.iterrows():
        body_rows.append([str(row.get(col, "") or "") for col in columns])

    # Ширины: под «бланковый» вид, но с читаемым текстом.
    col_widths = [9 * mm, 34 * mm, 46 * mm, 24 * mm] + [7.2 * mm for _ in day_cols] + [12 * mm, 14 * mm]
    table = Table(body_rows, colWidths=col_widths, repeatRows=1)
    table_style = TableStyle(
        [
            ("GRID", (0, 0), (-1, -1), 0.35, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dce6f2")),
            ("FONTNAME", (0, 0), (-1, 0), font_name),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTNAME", (0, 1), (-1, -1), font_name),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("ALIGN", (4, 1), (-1, -1), "CENTER"),
            ("ALIGN", (0, 1), (3, -1), "LEFT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fbff")]),
        ]
    )

    # Подсветим строки-итоги и строки "часы".
    for i, row in enumerate(body_rows[1:], start=1):
        role_val = row[3].strip().lower()
        num_val = row[0].strip().upper()
        if role_val == "часы":
            table_style.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#f2f2f2"))
            table_style.add("FONTSIZE", (0, i), (-1, i), 6.6)
        if num_val == "ИТОГО":
            table_style.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#e8f0fe"))
            table_style.add("FONTNAME", (0, i), (-1, i), font_name)

    table.setStyle(table_style)
    story.append(table)
    doc.build(story)
