#!/usr/bin/env python3
from __future__ import annotations

import argparse
import calendar
import sys
from collections import Counter
from datetime import date, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.work_rules import (
    extract_t13_records,
    minor_daily_cap,
    normalize_name,
    normalize_tab,
    parse_annual_payroll_register,
    parse_personnel_events,
)


SEVERITY = {
    "CRITICAL": 0,
    "ERROR": 1,
    "WARNING": 2,
    "INFO": 3,
}


def _employee_intervals(
    personnel: dict[str, list[tuple[date | None, date | None]]],
    employee: str,
) -> list[tuple[date | None, date | None]]:
    return personnel.get(normalize_name(employee), [])


def _is_employed(
    intervals: list[tuple[date | None, date | None]],
    current: date,
) -> bool:
    if not intervals:
        return True
    return any(
        (start is None or current >= start) and (finish is None or current <= finish)
        for start, finish in intervals
    )


def _register_lookup(register: dict[str, dict[str, Any]]) -> tuple[dict[str, Any], dict[str, Any]]:
    by_tab: dict[str, Any] = {}
    by_name: dict[str, Any] = {}
    for key, rec in register.items():
        tab = normalize_tab(rec.get("tab_number") or key)
        name = normalize_name(rec.get("employee", ""))
        if tab:
            by_tab[tab] = rec
        if name:
            by_name[name] = rec
    return by_tab, by_name


def _issue(
    issues: list[dict[str, Any]],
    *,
    year: int,
    month: int,
    issue_type: str,
    severity: str,
    employee: str = "",
    tab_number: str = "",
    day: int | None = None,
    t13_value: Any = "",
    payroll_value: Any = "",
    details: str = "",
) -> None:
    issues.append(
        {
            "year": year,
            "month": month,
            "severity": severity,
            "issue_type": issue_type,
            "tab_number": tab_number,
            "employee": employee,
            "day": day,
            "t13_value": t13_value,
            "payroll_value": payroll_value,
            "details": details,
        }
    )


def audit_month(
    *,
    year: int,
    month: int,
    filename: str,
    workbook_bytes: bytes,
    register_bytes: bytes,
    register_filename: str,
    personnel: dict[str, list[tuple[date | None, date | None]]],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    wb = load_workbook(BytesIO(workbook_bytes), data_only=True, read_only=True)
    ws = wb["Т-13 общий"] if "Т-13 общий" in wb.sheetnames else wb[wb.sheetnames[0]]
    period, t13 = extract_t13_records(ws, filename=filename, expected_period=(year, month))
    register = parse_annual_payroll_register(
        register_bytes,
        year=year,
        month=month,
        filename=register_filename,
    )
    register_by_tab, register_by_name = _register_lookup(register)
    matched_register_keys: set[str] = set()

    if period and period != (year, month):
        _issue(
            issues,
            year=year,
            month=month,
            issue_type="PERIOD_MISMATCH",
            severity="CRITICAL",
            t13_value=f"{period[1]:02d}.{period[0]}",
            payroll_value=f"{month:02d}.{year}",
            details="Отчетный период Т-13 не совпадает с месяцем архива.",
        )

    for tab, rec in t13.items():
        employee = rec["employee"]
        register_rec = register_by_tab.get(tab) or register_by_name.get(normalize_name(employee))
        if not register_rec:
            _issue(
                issues,
                year=year,
                month=month,
                issue_type="T13_EMPLOYEE_NOT_IN_PAYROLL",
                severity="WARNING",
                employee=employee,
                tab_number=tab,
                details="Сотрудник есть в Т-13, но не найден в расчетной ведомости месяца.",
            )
            continue

        register_key = normalize_tab(register_rec.get("tab_number")) or normalize_name(register_rec.get("employee"))
        matched_register_keys.add(register_key)
        work_hours = rec["work_hours"]
        actual_days = len(work_hours)
        actual_hours = round(sum(float(value or 0) for value in work_hours.values()), 2)
        expected_days = register_rec.get("register_days")
        expected_hours = register_rec.get("register_hours")

        if expected_days is not None and actual_days != int(expected_days):
            _issue(
                issues,
                year=year,
                month=month,
                issue_type="SHIFT_TOTAL_MISMATCH",
                severity="ERROR",
                employee=employee,
                tab_number=tab,
                t13_value=actual_days,
                payroll_value=int(expected_days),
                details="Количество явок в Т-13 не равно рабочим дням расчетной ведомости.",
            )
        if expected_hours is not None and abs(actual_hours - float(expected_hours)) > 0.02:
            _issue(
                issues,
                year=year,
                month=month,
                issue_type="HOURS_TOTAL_MISMATCH",
                severity="ERROR",
                employee=employee,
                tab_number=tab,
                t13_value=actual_hours,
                payroll_value=float(expected_hours),
                details="Сумма часов Т-13 не равна часам расчетной ведомости.",
            )

        intervals = _employee_intervals(personnel, employee)
        sick_days = set(register_rec.get("sick_days", set()))
        vacation_days = set(register_rec.get("vacation_days", set()))
        birth_date = register_rec.get("birth_date")
        weekly_hours: dict[tuple[int, int], float] = {}

        for day, hours in work_hours.items():
            current = date(year, month, day)
            if not _is_employed(intervals, current):
                _issue(
                    issues,
                    year=year,
                    month=month,
                    issue_type="WORK_OUTSIDE_EMPLOYMENT",
                    severity="CRITICAL",
                    employee=employee,
                    tab_number=tab,
                    day=day,
                    t13_value=hours,
                    details="Явка стоит вне интервала приема и увольнения.",
                )
            if day in sick_days:
                _issue(
                    issues,
                    year=year,
                    month=month,
                    issue_type="WORK_DURING_SICK_LEAVE",
                    severity="CRITICAL",
                    employee=employee,
                    tab_number=tab,
                    day=day,
                    t13_value=hours,
                    details="Явка пересекается с больничным из расчетной ведомости.",
                )
            if day in vacation_days:
                _issue(
                    issues,
                    year=year,
                    month=month,
                    issue_type="WORK_DURING_VACATION",
                    severity="CRITICAL",
                    employee=employee,
                    tab_number=tab,
                    day=day,
                    t13_value=hours,
                    details="Явка пересекается с отпуском из расчетной ведомости.",
                )
            if float(hours or 0) <= 0:
                _issue(
                    issues,
                    year=year,
                    month=month,
                    issue_type="WORK_CODE_WITHOUT_HOURS",
                    severity="ERROR",
                    employee=employee,
                    tab_number=tab,
                    day=day,
                    t13_value=hours,
                    details="Есть код Я, но часы не заполнены.",
                )

            cap = minor_daily_cap(birth_date, current)
            if cap < 13.0 and float(hours or 0) > cap + 0.001:
                _issue(
                    issues,
                    year=year,
                    month=month,
                    issue_type="MINOR_DAILY_LIMIT",
                    severity="CRITICAL",
                    employee=employee,
                    tab_number=tab,
                    day=day,
                    t13_value=hours,
                    payroll_value=cap,
                    details="Превышен дневной лимит несовершеннолетнего по принятому правилу.",
                )
            iso_year, iso_week, _ = current.isocalendar()
            weekly_hours[(iso_year, iso_week)] = weekly_hours.get((iso_year, iso_week), 0.0) + float(hours or 0)

        if birth_date:
            month_date = date(year, month, 15)
            age = month_date.year - birth_date.year - (
                (month_date.month, month_date.day) < (birth_date.month, birth_date.day)
            )
            if age < 18:
                study_month = month in {1, 2, 3, 4, 5, 9, 10, 11, 12}
                weekly_cap = (12.0 if age < 16 else 17.5) if study_month else (24.0 if age < 16 else 35.0)
                for (iso_year, iso_week), hours in weekly_hours.items():
                    if hours > weekly_cap + 0.001:
                        _issue(
                            issues,
                            year=year,
                            month=month,
                            issue_type="MINOR_WEEKLY_LIMIT",
                            severity="CRITICAL",
                            employee=employee,
                            tab_number=tab,
                            t13_value=round(hours, 2),
                            payroll_value=weekly_cap,
                            details=f"Превышен недельный лимит, ISO-неделя {iso_week}/{iso_year}.",
                        )

    for key, rec in register.items():
        register_key = normalize_tab(rec.get("tab_number") or key) or normalize_name(rec.get("employee"))
        expected_days = rec.get("register_days")
        expected_hours = rec.get("register_hours")
        if register_key in matched_register_keys or not expected_days or not expected_hours:
            continue
        _issue(
            issues,
            year=year,
            month=month,
            issue_type="PAYROLL_EMPLOYEE_NOT_IN_T13",
            severity="ERROR",
            employee=rec.get("employee", ""),
            tab_number=normalize_tab(rec.get("tab_number")),
            payroll_value=f"{expected_days} дн. / {expected_hours:g} ч.",
            details="В расчетной ведомости есть рабочие дни и часы, но сотрудник отсутствует в общем Т-13.",
        )

    counts = Counter(item["severity"] for item in issues)
    coverage = {
        "year": year,
        "month": month,
        "source_file": filename,
        "t13_employees": len(t13),
        "payroll_employees": len(register),
        "critical": counts["CRITICAL"],
        "errors": counts["ERROR"],
        "warnings": counts["WARNING"],
        "status": "FAIL" if counts["CRITICAL"] or counts["ERROR"] else ("WARN" if counts["WARNING"] else "OK"),
    }
    return issues, coverage


def _write_table(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, header in enumerate(headers, start=1):
        max_len = max([len(str(header))] + [len(str(row.get(header, ""))) for row in rows[:1000]])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 55)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_report(output: Path, issues: list[dict[str, Any]], coverage: list[dict[str, Any]]) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Итоги"
    total_counts = Counter(issue["severity"] for issue in issues)
    summary_rows = [
        {"Показатель": "Проверено месяцев", "Значение": len(coverage)},
        {"Показатель": "Критические нарушения", "Значение": total_counts["CRITICAL"]},
        {"Показатель": "Ошибки сверки", "Значение": total_counts["ERROR"]},
        {"Показатель": "Предупреждения", "Значение": total_counts["WARNING"]},
        {
            "Показатель": "Месяцев без критических нарушений и ошибок",
            "Значение": sum(item["status"] in {"OK", "WARN"} for item in coverage),
        },
    ]
    _write_table(summary, ["Показатель", "Значение"], summary_rows)

    coverage_ws = wb.create_sheet("Помесячная проверка")
    coverage_columns = [
        ("Год", "year"),
        ("Месяц", "month"),
        ("Файл Т-13", "source_file"),
        ("Сотрудников в Т-13", "t13_employees"),
        ("Сотрудников в ведомости", "payroll_employees"),
        ("Критические", "critical"),
        ("Ошибки", "errors"),
        ("Предупреждения", "warnings"),
        ("Статус", "status"),
    ]
    coverage_rows = [
        {title: row.get(key, "") for title, key in coverage_columns}
        for row in sorted(coverage, key=lambda row: (row["year"], row["month"]))
    ]
    _write_table(coverage_ws, [title for title, _ in coverage_columns], coverage_rows)

    issues_ws = wb.create_sheet("Нарушения")
    issue_columns = [
        ("Год", "year"),
        ("Месяц", "month"),
        ("Критичность", "severity"),
        ("Тип нарушения", "issue_type"),
        ("Табельный номер", "tab_number"),
        ("Сотрудник", "employee"),
        ("День", "day"),
        ("Значение Т-13", "t13_value"),
        ("Расчетная ведомость / лимит", "payroll_value"),
        ("Описание", "details"),
    ]
    issues_sorted = sorted(
        issues,
        key=lambda row: (
            row["year"],
            row["month"],
            SEVERITY[row["severity"]],
            row["employee"],
            row["day"] or 0,
        ),
    )
    issue_rows = [
        {title: row.get(key, "") for title, key in issue_columns}
        for row in issues_sorted
    ]
    _write_table(issues_ws, [title for title, _ in issue_columns], issue_rows)
    severity_fills = {
        "CRITICAL": "F4CCCC",
        "ERROR": "FCE5CD",
        "WARNING": "FFF2CC",
        "INFO": "D9EAD3",
    }
    for row_idx in range(2, issues_ws.max_row + 1):
        fill = PatternFill("solid", fgColor=severity_fills.get(issues_ws.cell(row_idx, 3).value, "FFFFFF"))
        for cell in issues_ws[row_idx]:
            cell.fill = fill

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> None:
    parser = argparse.ArgumentParser(description="Audit historical T-13 archives against payroll and HR rules.")
    parser.add_argument("--zip", action="append", required=True, help="Path to annual ZIP archive.")
    parser.add_argument("--register", action="append", required=True, help="YEAR=/path/to/annual register.xlsx")
    parser.add_argument("--personnel", required=True, help="Personnel hire/dismiss .xls file")
    parser.add_argument("--output", required=True, help="Output .xlsx report")
    args = parser.parse_args()

    registers: dict[int, tuple[str, bytes]] = {}
    for item in args.register:
        year_text, path_text = item.split("=", 1)
        path = Path(path_text)
        registers[int(year_text)] = (path.name, path.read_bytes())

    personnel_path = Path(args.personnel)
    personnel = parse_personnel_events(personnel_path.read_bytes(), personnel_path.name)
    all_issues: list[dict[str, Any]] = []
    all_coverage: list[dict[str, Any]] = []

    for zip_text in args.zip:
        zip_path = Path(zip_text)
        with ZipFile(zip_path) as archive:
            for filename in sorted(archive.namelist()):
                if not filename.lower().endswith(".xlsx"):
                    continue
                workbook_bytes = archive.read(filename)
                wb = load_workbook(BytesIO(workbook_bytes), data_only=True, read_only=True)
                ws = wb["Т-13 общий"] if "Т-13 общий" in wb.sheetnames else wb[wb.sheetnames[0]]
                period, _ = extract_t13_records(ws, filename=filename)
                if not period:
                    continue
                year, month = period
                if year not in registers:
                    continue
                register_filename, register_bytes = registers[year]
                issues, coverage = audit_month(
                    year=year,
                    month=month,
                    filename=filename,
                    workbook_bytes=workbook_bytes,
                    register_bytes=register_bytes,
                    register_filename=register_filename,
                    personnel=personnel,
                )
                all_issues.extend(issues)
                all_coverage.append(coverage)

    write_report(Path(args.output), all_issues, all_coverage)
    counts = Counter(issue["severity"] for issue in all_issues)
    print(
        f"months={len(all_coverage)} issues={len(all_issues)} "
        f"critical={counts['CRITICAL']} errors={counts['ERROR']} warnings={counts['WARNING']}"
    )


if __name__ == "__main__":
    main()
